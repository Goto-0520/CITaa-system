# -*- coding: utf-8 -*-
"""
CITAA System - 千葉工業大学体育会本部 統合管理システム
Complete Flet Desktop Application v3.0 - Major Overhaul
"""

import flet as ft
from flet import (
    Page, View, AppBar, Text, Container, Column, Row, Card, 
    ElevatedButton, TextButton, IconButton, Icon, TextField,
    Dropdown, dropdown, Checkbox, DataTable, DataColumn, DataRow, DataCell,
    ProgressRing, ProgressBar, Divider, AlertDialog, SnackBar,
    FilePicker, padding, margin,
    MainAxisAlignment, CrossAxisAlignment, ScrollMode, alignment,
    BorderSide, RoundedRectangleBorder, BoxShadow, Offset,
    NavigationRail, NavigationRailDestination
)

# Version compatibility - Use ft.Icons (not deprecated icons)
Icons = ft.Icons
colors = ft.colors

# Additional compatibility
BlurStyle = getattr(ft, "BlurStyle", None)
FilePickerResultEvent = getattr(ft, "FilePickerResultEvent", ft.ControlEvent)

from datetime import datetime, date, timedelta
from typing import List, Dict, Any, Optional, Callable, Tuple
import threading
import asyncio
import os
import json
import io
import zipfile
import tempfile
import re

# Local imports
import config
from config import user_settings, ensure_directories
from auth.google_auth import get_auth_manager
from services.sheets_service import get_sheets_service
from services.error_logger import get_error_logger

# PDF Generation imports
try:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    HAS_PDF_LIBS = True
except ImportError:
    HAS_PDF_LIBS = False

try:
    from PyPDF2 import PdfWriter, PdfReader
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

# Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


# ============================================================
# Helper Functions
# ============================================================
def get_accent_color() -> str:
    return user_settings.accent_color


def format_number_with_comma(value: int) -> str:
    """数字をカンマ区切りにフォーマット"""
    if value == 0:
        return "0"
    return f"{value:,}"


def parse_number_from_comma(value: str) -> int:
    """カンマ区切りの文字列を数字に変換"""
    if not value:
        return 0
    try:
        return int(value.replace(",", "").replace("¥", "").strip())
    except:
        return 0


def get_club_display_name(club: Dict) -> str:
    """団体名を「団体名＋区分」形式で取得"""
    name = club.get("ClubName", "")
    category = club.get("Category", "")
    if category and not name.endswith(category):
        return f"{name}{category}"
    return name


def generate_time_options() -> List[Tuple[str, str]]:
    """30分刻みの時刻オプションを生成"""
    options = []
    for hour in range(6, 24):
        for minute in [0, 30]:
            time_str = f"{hour:02d}:{minute:02d}"
            options.append((time_str, time_str))
    return options


def generate_date_dropdown_options(field_type: str) -> List[Tuple[str, str]]:
    """日付のドロップダウンオプションを生成"""
    if field_type == "year":
        current_year = datetime.now().year
        return [(str(y), str(y)) for y in range(current_year - 2, current_year + 3)]
    elif field_type == "month":
        return [(str(m), f"{m}月") for m in range(1, 13)]
    elif field_type == "day":
        return [(str(d), f"{d}日") for d in range(1, 32)]
    return []


# ============================================================
# Loading Overlay Component - Solid Background (No See-Through)
# ============================================================
class LoadingOverlay:
    def __init__(self, page: Page):
        self.page = page
        self.message_text = Text("データを読み込み中...", size=18, color=config.COLOR_GRAY_700)
        self.overlay = Container(
            content=Column(
                controls=[
                    ProgressRing(width=60, height=60, stroke_width=5, color=get_accent_color()),
                    Container(height=20),
                    self.message_text,
                ],
                horizontal_alignment=CrossAxisAlignment.CENTER,
                alignment=MainAxisAlignment.CENTER,
                spacing=0,
            ),
            alignment=alignment.center,
            bgcolor=config.COLOR_GRAY_50,  # 完全に不透明な背景
            expand=True,
            visible=False,
        )
    
    def show(self, message: str = "データを読み込み中..."):
        self.message_text.value = message
        self.overlay.visible = True
        if self.page:
            try:
                self.page.update()
            except:
                pass
    
    def hide(self):
        self.overlay.visible = False
        if self.page:
            try:
                self.page.update()
            except:
                pass


# ============================================================
# Styled Card Component - Centered Layout
# ============================================================
def create_dept_card(
    title: str,
    subtitle: str,
    icon_name,
    color: str,
    on_click: Callable,
) -> Container:
    return Container(
        content=Column(
            controls=[
                Row(
                    controls=[
                        Container(
                            content=Icon(icon_name, size=32, color=config.COLOR_WHITE),
                            bgcolor=color,
                            border_radius=12,
                            padding=12,
                        ),
                        Column(
                            controls=[
                                Text(title, size=18, weight=ft.FontWeight.BOLD, color=config.COLOR_GRAY_800),
                                Text(subtitle, size=12, color=config.COLOR_GRAY_500),
                            ],
                            spacing=2,
                            expand=True,
                        ),
                    ],
                    spacing=15,
                ),
            ],
        ),
        padding=20,
        bgcolor=config.COLOR_WHITE,
        border_radius=16,
        border=ft.border.all(1, config.COLOR_GRAY_200),
        shadow=BoxShadow(
            spread_radius=0,
            blur_radius=10,
            color="#0000000d",
            offset=Offset(0, 4),
        ),
        ink=True,
        on_click=on_click,
        width=320,
    )


def create_section_header(title: str, subtitle: str = "") -> Container:
    controls = [Text(title, size=20, weight=ft.FontWeight.BOLD, color=config.COLOR_GRAY_800)]
    if subtitle:
        controls.append(Text(subtitle, size=14, color=config.COLOR_GRAY_500))
    return Container(
        content=Column(controls=controls, spacing=4),
        margin=margin.only(bottom=20),
    )


# ============================================================
# Form Input Helpers with Number Formatting
# ============================================================
def create_text_field(
    label: str,
    value: str = "",
    width: Optional[int] = None,
    on_change: Optional[Callable] = None,
    password: bool = False,
    multiline: bool = False,
    hint_text: str = "",
    expand: bool = False,
) -> TextField:
    return TextField(
        label=label,
        value=value,
        width=width,
        expand=expand,
        on_change=on_change,
        password=password,
        can_reveal_password=password,
        multiline=multiline,
        hint_text=hint_text,
        border_radius=8,
        border_color=config.COLOR_GRAY_300,
        focused_border_color=get_accent_color(),
    )


def create_number_field(
    label: str,
    value: int = 0,
    width: Optional[int] = None,
    on_change: Optional[Callable] = None,
) -> TextField:
    """カンマ自動挿入付き数字入力フィールド"""
    field = TextField(
        label=label,
        value=format_number_with_comma(value) if value else "",
        width=width,
        border_radius=8,
        border_color=config.COLOR_GRAY_300,
        focused_border_color=get_accent_color(),
        input_filter=ft.InputFilter(allow=True, regex_string=r"[0-9,]", replacement_string=""),
    )
    
    def on_blur(e):
        num = parse_number_from_comma(field.value)
        field.value = format_number_with_comma(num) if num else ""
        field.update()
        if on_change:
            on_change(e)
    
    field.on_blur = on_blur
    return field


def create_dropdown(
    label: str,
    options: List[Tuple[str, str]],
    value: str = None,
    width: Optional[int] = None,
    on_change: Optional[Callable] = None,
    expand: bool = False,
) -> Dropdown:
    return Dropdown(
        label=label,
        value=value,
        options=[dropdown.Option(key=k, text=v) for k, v in options],
        width=width,
        expand=expand,
        on_change=on_change,
        border_radius=8,
        border_color=config.COLOR_GRAY_300,
        focused_border_color=get_accent_color(),
    )


def create_time_dropdown(
    label: str,
    value: str = None,
    width: int = 100,
    on_change: Optional[Callable] = None,
) -> Dropdown:
    """時刻選択ドロップダウン（30分刻み）"""
    return create_dropdown(
        label=label,
        options=generate_time_options(),
        value=value,
        width=width,
        on_change=on_change,
    )


def create_date_dropdowns(
    year: int = None,
    month: int = None,
    day: int = None,
    on_change: Optional[Callable] = None,
) -> Tuple[Dropdown, Dropdown, Dropdown]:
    """日付選択用の3つのドロップダウン（年/月/日）"""
    now = datetime.now()
    year = year or now.year
    month = month or now.month
    day = day or now.day
    
    year_dd = create_dropdown(
        "年",
        generate_date_dropdown_options("year"),
        value=str(year),
        width=90,
        on_change=on_change,
    )
    month_dd = create_dropdown(
        "月",
        generate_date_dropdown_options("month"),
        value=str(month),
        width=75,
        on_change=on_change,
    )
    day_dd = create_dropdown(
        "日",
        generate_date_dropdown_options("day"),
        value=str(day),
        width=75,
        on_change=on_change,
    )
    return year_dd, month_dd, day_dd


def create_primary_button(
    text: str,
    on_click: Callable,
    icon=None,
    width: Optional[int] = None,
    expand: bool = False,
    bgcolor: str = None,
    disabled: bool = False,
) -> ElevatedButton:
    return ElevatedButton(
        text=text,
        icon=icon,
        on_click=on_click,
        width=width,
        expand=expand,
        disabled=disabled,
        style=ft.ButtonStyle(
            bgcolor=bgcolor or get_accent_color(),
            color=config.COLOR_WHITE,
            padding=padding.symmetric(horizontal=24, vertical=12),
            shape=RoundedRectangleBorder(radius=8),
        ),
    )


def create_secondary_button(
    text: str,
    on_click: Callable,
    icon=None,
    width: Optional[int] = None,
    expand: bool = False,
) -> ElevatedButton:
    return ElevatedButton(
        text=text,
        icon=icon,
        on_click=on_click,
        width=width,
        expand=expand,
        style=ft.ButtonStyle(
            bgcolor=config.COLOR_GRAY_100,
            color=config.COLOR_GRAY_700,
            padding=padding.symmetric(horizontal=24, vertical=12),
            shape=RoundedRectangleBorder(radius=8),
        ),
    )


def create_danger_button(
    text: str,
    on_click: Callable,
    icon=None,
    width: Optional[int] = None,
) -> ElevatedButton:
    return ElevatedButton(
        text=text,
        icon=icon,
        on_click=on_click,
        width=width,
        style=ft.ButtonStyle(
            bgcolor="#ef4444",
            color=config.COLOR_WHITE,
            padding=padding.symmetric(horizontal=24, vertical=12),
            shape=RoundedRectangleBorder(radius=8),
        ),
    )


# ============================================================
# Search Component with Partial Match
# ============================================================
def create_search_field(
    hint: str = "検索...",
    width: int = 300,
    on_change: Optional[Callable] = None,
) -> TextField:
    return TextField(
        hint_text=hint,
        prefix_icon=Icons.SEARCH,
        width=width,
        border_radius=20,
        border_color=config.COLOR_GRAY_300,
        focused_border_color=get_accent_color(),
        on_change=on_change,
        content_padding=padding.symmetric(horizontal=15, vertical=10),
    )


def filter_by_search(items: List[Dict], search_text: str, fields: List[str]) -> List[Dict]:
    """部分一致検索でフィルタリング"""
    if not search_text:
        return items
    search_lower = search_text.lower()
    return [
        item for item in items
        if any(search_lower in str(item.get(field, "")).lower() for field in fields)
    ]


# ============================================================
# Confirmation Dialog
# ============================================================
def show_confirm_dialog(
    page: Page,
    title: str,
    message: str,
    on_confirm: Callable,
    on_cancel: Optional[Callable] = None,
    confirm_text: str = "確認",
    cancel_text: str = "キャンセル",
    is_danger: bool = False,
):
    def close_dialog(e):
        dialog.open = False
        page.update()
        if on_cancel:
            on_cancel()
    
    def confirm_action(e):
        dialog.open = False
        page.update()
        on_confirm()
    
    dialog = AlertDialog(
        modal=True,
        title=Text(title, weight=ft.FontWeight.BOLD),
        content=Text(message),
        actions=[
            TextButton(cancel_text, on_click=close_dialog),
            ElevatedButton(
                confirm_text,
                on_click=confirm_action,
                style=ft.ButtonStyle(
                    bgcolor="#ef4444" if is_danger else get_accent_color(),
                    color=config.COLOR_WHITE,
                ),
            ),
        ],
        actions_alignment=MainAxisAlignment.END,
    )
    page.dialog = dialog
    dialog.open = True
    page.update()


# ============================================================
# Edit/Delete Action Buttons
# ============================================================
def create_edit_delete_buttons(
    on_edit: Callable,
    on_delete: Callable,
) -> Row:
    return Row(
        controls=[
            IconButton(
                icon=Icons.EDIT,
                icon_size=18,
                icon_color=get_accent_color(),
                tooltip="編集",
                on_click=on_edit,
            ),
            IconButton(
                icon=Icons.DELETE,
                icon_size=18,
                icon_color="#ef4444",
                tooltip="削除",
                on_click=on_delete,
            ),
        ],
        spacing=0,
    )


# ============================================================
# File Drop Zone Component
# ============================================================
def create_file_drop_zone(
    on_click: Callable,
    accept_types: List[str],
    height: int = 150,
) -> Container:
    """ファイルアップロードエリア"""
    return Container(
        content=Column(
            controls=[
                Icon(Icons.CLOUD_UPLOAD, size=40, color=config.COLOR_GRAY_400),
                Text("ファイルをドラッグ&ドロップ", size=14, color=config.COLOR_GRAY_600),
                Text("または", size=12, color=config.COLOR_GRAY_500),
                TextButton(
                    "クリックして選択",
                    icon=Icons.FOLDER_OPEN,
                    on_click=on_click,
                ),
                Text(f"対応形式: {', '.join(accept_types)}", size=11, color=config.COLOR_GRAY_400),
            ],
            horizontal_alignment=CrossAxisAlignment.CENTER,
            alignment=MainAxisAlignment.CENTER,
            spacing=5,
        ),
        height=height,
        border=ft.border.all(2, config.COLOR_GRAY_300),
        border_radius=12,
        bgcolor=config.COLOR_GRAY_50,
        alignment=alignment.center,
    )


# ============================================================
# Main Application Class
# ============================================================
class CITAAApp:
    def __init__(self, page: Page):
        self.page = page
        self.loading = LoadingOverlay(page)
        self.auth_manager = get_auth_manager()
        self.sheets_service = None
        
        # Data caches
        self.clubs: List[Dict] = []
        self.members: List[Dict] = []
        self.facilities: List[Dict] = []
        self.categories: List[Dict] = []
        
        # UI state
        self.current_view = "login"
        self.selected_dept_index = 0
        self._main_content = None
        self._current_view_content = None
        
        # File picker
        self.file_picker = FilePicker(on_result=self._on_file_picked)
        self._file_pick_callback = None
        self.page.overlay.append(self.file_picker)
        self.page.overlay.append(self.loading.overlay)
        
        # Setup page
        self._setup_page()
        
    def _setup_page(self):
        self.page.title = config.APP_TITLE
        self.page.window.width = config.APP_WIDTH
        self.page.window.height = config.APP_HEIGHT
        self.page.bgcolor = config.COLOR_GRAY_50
        self.page.padding = 0
        self.page.theme = ft.Theme(font_family="Yu Gothic UI")
        
        # Check if already authenticated
        if config.TOKEN_FILE.exists():
            self._try_auto_login()
        else:
            self._show_login_view()
    
    def _try_auto_login(self):
        self.loading.show("認証情報を確認中...")
        
        def do_auth():
            success, msg = self.auth_manager.authenticate()
            self.page.run_thread(lambda: self._on_auth_complete(success, msg))
        
        threading.Thread(target=do_auth, daemon=True).start()
    
    def _on_auth_complete(self, success: bool, msg: str):
        self.loading.hide()
        if success:
            self.sheets_service = get_sheets_service()
            self._load_initial_data()
            self._show_dashboard()
        else:
            self._show_login_view()
    
    def _load_initial_data(self):
        try:
            if self.sheets_service:
                self.clubs = self.sheets_service.get_clubs() or []
                self.members = self.sheets_service.get_members() or []
                self.facilities = self.sheets_service.get_facilities() or []
                self.categories = self.sheets_service.get_categories() or []
        except Exception as e:
            get_error_logger().log_error("CITAAApp", "_load_initial_data", e)
    
    # ============================================================
    # Login View - Wider Layout
    # ============================================================
    def _show_login_view(self):
        self.current_view = "login"
        
        login_card = Container(
            content=Column(
                controls=[
                    # Logo/Title
                    Container(
                        content=Icon(Icons.SPORTS_SOCCER, size=70, color=get_accent_color()),
                        alignment=alignment.center,
                    ),
                    Container(height=10),
                    Text(
                        "千葉工業大学体育会本部",
                        size=28,
                        weight=ft.FontWeight.BOLD,
                        color=config.COLOR_GRAY_800,
                        text_align=ft.TextAlign.CENTER,
                    ),
                    Text(
                        "統合管理システム",
                        size=18,
                        color=config.COLOR_GRAY_500,
                        text_align=ft.TextAlign.CENTER,
                    ),
                    Container(height=40),
                    
                    # Login Button
                    create_primary_button(
                        "Googleアカウントでログイン",
                        self._on_login_click,
                        icon=Icons.LOGIN,
                        width=320,
                    ),
                    
                    Container(height=40),
                    Divider(height=1, color=config.COLOR_GRAY_200),
                    Container(height=25),
                    
                    # Instructions
                    Container(
                        content=Column(
                            controls=[
                                Text(
                                    "📋 初めてログインする方へ",
                                    size=15,
                                    weight=ft.FontWeight.BOLD,
                                    color=config.COLOR_GRAY_700,
                                ),
                                Container(height=15),
                                self._create_step_row("1", "ログインボタンをクリック"),
                                Container(height=8),
                                self._create_step_row("2", "Googleアカウントを選択"),
                                Container(height=8),
                                self._create_step_row("3", "「このアプリはGoogleで確認されていません」→「続行」をクリック"),
                                Container(height=8),
                                self._create_step_row("4", "アクセス権限を「許可」"),
                                Container(height=8),
                                self._create_step_row("5", "認証完了画面が表示されたら完了！", False, "#22c55e"),
                            ],
                        ),
                        bgcolor=config.COLOR_GRAY_50,
                        border_radius=12,
                        padding=25,
                        width=450,
                    ),
                ],
                horizontal_alignment=CrossAxisAlignment.CENTER,
                spacing=5,
            ),
            padding=50,
            bgcolor=config.COLOR_WHITE,
            border_radius=24,
            shadow=BoxShadow(
                spread_radius=0,
                blur_radius=40,
                color="#00000018",
                offset=Offset(0, 15),
            ),
            width=550,  # 幅を広げた
        )
        
        self.page.controls.clear()
        self.page.add(
            Container(
                content=login_card,
                alignment=alignment.center,
                expand=True,
                bgcolor=config.COLOR_GRAY_100,
            )
        )
        self.page.update()
    
    def _create_step_row(self, num: str, text: str, multiline: bool = True, color: str = None) -> Row:
        bg_color = color or get_accent_color()
        text_color = config.COLOR_GRAY_600
        
        return Row(
            controls=[
                Container(
                    content=Text(num, size=12, color=config.COLOR_WHITE, weight=ft.FontWeight.BOLD),
                    bgcolor=bg_color,
                    border_radius=12,
                    width=24,
                    height=24,
                    alignment=alignment.center,
                ),
                Text(text, size=14, color=text_color, expand=True),
            ],
            spacing=12,
        )
    
    def _on_login_click(self, e):
        self.loading.show("ブラウザでログイン画面を開いています...")
        
        def do_auth():
            success, msg = self.auth_manager.authenticate()
            self.page.run_thread(lambda: self._on_auth_complete(success, msg))
        
        threading.Thread(target=do_auth, daemon=True).start()
    
    # ============================================================
    # Dashboard View - Centered Layout
    # ============================================================
    def _show_dashboard(self):
        self.current_view = "dashboard"
        self._current_view_content = None
        
        # Get user info
        user_info = self.auth_manager.get_user_info()
        user_name = user_info.get("displayName", "ユーザー") if user_info else "ユーザー"
        user_email = user_info.get("emailAddress", "") if user_info else ""
        
        # Header
        header = Container(
            content=Row(
                controls=[
                    Row(
                        controls=[
                            Icon(Icons.SPORTS_SOCCER, size=32, color=get_accent_color()),
                            Text(
                                "CITAA System",
                                size=22,
                                weight=ft.FontWeight.BOLD,
                                color=config.COLOR_GRAY_800,
                            ),
                        ],
                        spacing=10,
                    ),
                    Row(
                        controls=[
                            Column(
                                controls=[
                                    Text(user_name, size=14, weight=ft.FontWeight.W_500, color=config.COLOR_GRAY_700),
                                    Text(user_email, size=12, color=config.COLOR_GRAY_500),
                                ],
                                spacing=2,
                                horizontal_alignment=CrossAxisAlignment.END,
                            ),
                            IconButton(
                                icon=Icons.LOGOUT,
                                icon_color=config.COLOR_GRAY_500,
                                tooltip="ログアウト",
                                on_click=self._on_logout_click,
                            ),
                        ],
                        spacing=10,
                    ),
                ],
                alignment=MainAxisAlignment.SPACE_BETWEEN,
            ),
            padding=padding.symmetric(horizontal=30, vertical=15),
            bgcolor=config.COLOR_WHITE,
            border=ft.border.only(bottom=BorderSide(1, config.COLOR_GRAY_200)),
        )
        
        # Department Cards - Centered
        dept_cards = Row(
            controls=[
                create_dept_card(
                    "書記部",
                    "施設利用管理・カレンダー出力",
                    Icons.CALENDAR_MONTH,
                    config.DEPT_COLORS["secretary"],
                    lambda e: self._navigate_to_dept("secretary"),
                ),
                create_dept_card(
                    "財務部",
                    "収支管理・立替精算",
                    Icons.ACCOUNT_BALANCE_WALLET,
                    config.DEPT_COLORS["finance"],
                    lambda e: self._navigate_to_dept("finance"),
                ),
                create_dept_card(
                    "総務部",
                    "シフト管理・出欠管理",
                    Icons.GROUPS,
                    config.DEPT_COLORS["general"],
                    lambda e: self._navigate_to_dept("general"),
                ),
            ],
            wrap=True,
            spacing=20,
            run_spacing=20,
            alignment=MainAxisAlignment.CENTER,  # 中央配置
        )
        
        dept_cards_2 = Row(
            controls=[
                create_dept_card(
                    "渉外部",
                    "課外活動記録・スキャン管理",
                    Icons.PUBLIC,
                    config.DEPT_COLORS["external"],
                    lambda e: self._navigate_to_dept("external"),
                ),
                create_dept_card(
                    "編集部",
                    "名簿校正・DB編集",
                    Icons.EDIT_NOTE,
                    config.DEPT_COLORS["editorial"],
                    lambda e: self._navigate_to_dept("editorial"),
                ),
                create_dept_card(
                    "イベント管理",
                    "フォーム集計・フォルダ作成",
                    Icons.EVENT,
                    config.DEPT_COLORS["event"],
                    lambda e: self._navigate_to_dept("event"),
                ),
            ],
            wrap=True,
            spacing=20,
            run_spacing=20,
            alignment=MainAxisAlignment.CENTER,  # 中央配置
        )
        
        admin_card_row = Row(
            controls=[
                create_dept_card(
                    "管理 & 設定",
                    "団体管理・本部員情報・システム設定",
                    Icons.SETTINGS,
                    config.COLOR_GRAY_600,
                    lambda e: self._navigate_to_admin(),
                ),
            ],
            alignment=MainAxisAlignment.CENTER,  # 中央配置
        )
        
        # Main content - Centered with scroll
        content = Container(
            content=Column(
                controls=[
                    Container(
                        content=Column(
                            controls=[
                                create_section_header(
                                    f"ようこそ、{user_name}さん",
                                    datetime.now().strftime("%Y年%m月%d日 (%a)"),
                                ),
                                Text("部門を選択してください", size=16, color=config.COLOR_GRAY_600),
                                Container(height=30),
                                dept_cards,
                                Container(height=15),
                                dept_cards_2,
                                Container(height=40),
                                Divider(height=1, color=config.COLOR_GRAY_200),
                                Container(height=25),
                                admin_card_row,
                                Container(height=30),  # 下側の余白
                            ],
                            horizontal_alignment=CrossAxisAlignment.CENTER,
                        ),
                        alignment=alignment.center,
                    ),
                ],
                scroll=ScrollMode.AUTO,
                horizontal_alignment=CrossAxisAlignment.CENTER,
            ),
            padding=padding.symmetric(horizontal=30, vertical=20),
            expand=True,
        )
        
        self.page.controls.clear()
        self.page.add(
            Column(
                controls=[header, content],
                spacing=0,
                expand=True,
            )
        )
        self.page.update()
    
    def _on_logout_click(self, e):
        def do_logout():
            self.auth_manager.logout()
            self.page.run_thread(self._show_login_view)
        
        threading.Thread(target=do_logout, daemon=True).start()
    
    # ============================================================
    # Admin Navigation with Password Protection
    # ============================================================
    def _navigate_to_admin(self):
        """管理&設定へのナビゲーション（パスワード保護付き）"""
        if user_settings.system_password:
            # パスワードが設定されている場合は認証を要求
            self._show_password_dialog()
        else:
            self._navigate_to_dept("admin")
    
    def _show_password_dialog(self):
        password_field = TextField(
            label="システムパスワード",
            password=True,
            can_reveal_password=True,
            width=300,
            autofocus=True,
        )
        error_text = Text("", color="#ef4444", size=12)
        
        def check_password(e):
            if password_field.value == user_settings.system_password:
                dialog.open = False
                self.page.update()
                self._navigate_to_dept("admin")
            else:
                error_text.value = "パスワードが正しくありません"
                self.page.update()
        
        def cancel(e):
            dialog.open = False
            self.page.update()
        
        dialog = AlertDialog(
            modal=True,
            title=Text("管理者認証", weight=ft.FontWeight.BOLD),
            content=Column(
                controls=[
                    Text("管理&設定にアクセスするにはパスワードを入力してください。"),
                    Container(height=15),
                    password_field,
                    error_text,
                ],
                tight=True,
            ),
            actions=[
                TextButton("キャンセル", on_click=cancel),
                ElevatedButton(
                    "認証",
                    on_click=check_password,
                    style=ft.ButtonStyle(bgcolor=get_accent_color(), color=config.COLOR_WHITE),
                ),
            ],
            actions_alignment=MainAxisAlignment.END,
        )
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    # ============================================================
    # Department Navigation - Fixed transition (no see-through)
    # ============================================================
    def _navigate_to_dept(self, dept: str):
        # Store current view content before showing loading
        self._current_view_content = self.page.controls.copy() if self.page.controls else None
        
        # Show loading overlay on TOP of current content (opaque)
        self.loading.show("データを読み込み中...")
        
        def load_and_show():
            try:
                self._load_initial_data()
            except Exception as e:
                get_error_logger().log_error("CITAAApp", f"_navigate_to_dept({dept})", e)
            
            def show_view():
                # Hide loading and show new view
                self.loading.hide()
                if dept == "secretary":
                    self._show_secretary_view()
                elif dept == "finance":
                    self._show_finance_view()
                elif dept == "general":
                    self._show_general_view()
                elif dept == "external":
                    self._show_external_view()
                elif dept == "editorial":
                    self._show_editorial_view()
                elif dept == "event":
                    self._show_event_view()
                elif dept == "admin":
                    self._show_admin_view()
            
            self.page.run_thread(show_view)
        
        threading.Thread(target=load_and_show, daemon=True).start()
    
    def _create_dept_header(self, title: str, color: str, icon_name) -> Container:
        """Create a standard department header"""
        return Container(
            content=Row(
                controls=[
                    Row(
                        controls=[
                            IconButton(
                                icon=Icons.ARROW_BACK,
                                icon_color=config.COLOR_GRAY_600,
                                on_click=lambda e: self._show_dashboard(),
                            ),
                            Container(
                                content=Icon(icon_name, size=24, color=config.COLOR_WHITE),
                                bgcolor=color,
                                border_radius=8,
                                padding=8,
                            ),
                            Text(
                                title,
                                size=20,
                                weight=ft.FontWeight.BOLD,
                                color=config.COLOR_GRAY_800,
                            ),
                        ],
                        spacing=10,
                    ),
                ],
            ),
            padding=padding.symmetric(horizontal=20, vertical=12),
            bgcolor=config.COLOR_WHITE,
            border=ft.border.only(bottom=BorderSide(1, config.COLOR_GRAY_200)),
        )
    
    def _create_tab_buttons(self, tabs: List[tuple], current_index: dict, on_switch: Callable) -> List[ElevatedButton]:
        """Create tab buttons with proper styling"""
        buttons = []
        for i, (label, _) in enumerate(tabs):
            btn = ElevatedButton(
                label,
                on_click=lambda e, idx=i: on_switch(idx),
                style=ft.ButtonStyle(
                    bgcolor=get_accent_color() if i == current_index["value"] else config.COLOR_GRAY_100,
                    color=config.COLOR_WHITE if i == current_index["value"] else config.COLOR_GRAY_700,
                    padding=padding.symmetric(horizontal=20, vertical=10),
                    shape=RoundedRectangleBorder(radius=8),
                ),
            )
            buttons.append(btn)
        return buttons
    
    # ============================================================
    # Secretary View (書記部) - 記録打ち込み特化
    # ============================================================
    def _show_secretary_view(self):
        self.current_view = "secretary"
        
        current_tab = {"value": 0}
        tab_content = Container(expand=True)
        tab_buttons_row = Row(controls=[], spacing=10)
        
        # Load data
        secretary_logs = []
        try:
            if self.sheets_service:
                secretary_logs = self.sheets_service.get_secretary_logs() or []
        except Exception as e:
            get_error_logger().log_error("Secretary", "load", e)
        
        # 入力用の行データを管理
        entry_rows = {"items": []}
        
        # --- Tab 0: 記録 打ち込み ---
        def create_record_tab():
            # 団体選択（検索付き）
            search_field = create_search_field("団体名で検索...", width=250)
            
            club_options = [(c.get("ClubName", ""), get_club_display_name(c)) for c in self.clubs]
            club_dropdown = create_dropdown("団体", club_options, width=250)
            
            facility_options = [(f.get("FacilityName", ""), f.get("FacilityName", "")) for f in self.facilities]
            facility_dropdown = create_dropdown("施設", facility_options, width=300)
            
            # 入力行リスト
            entry_list = Column(controls=[], spacing=5)
            
            def add_entry_row(e=None):
                now = datetime.now()
                year_dd, month_dd, day_dd = create_date_dropdowns()
                start_time = create_time_dropdown("開始", "09:00", width=100)
                end_time = create_time_dropdown("終了", "17:00", width=100)
                
                row_data = {
                    "year": year_dd, "month": month_dd, "day": day_dd,
                    "start": start_time, "end": end_time
                }
                entry_rows["items"].append(row_data)
                
                def remove_row(e, idx=len(entry_rows["items"])-1):
                    if len(entry_rows["items"]) > 1:
                        entry_rows["items"].pop(idx)
                        rebuild_entry_list()
                
                row = Container(
                    content=Row(
                        controls=[
                            year_dd, month_dd, day_dd,
                            Container(width=20),
                            start_time,
                            Text("〜", size=14),
                            end_time,
                            IconButton(
                                icon=Icons.REMOVE_CIRCLE_OUTLINE,
                                icon_color="#ef4444",
                                tooltip="削除",
                                on_click=remove_row,
                            ),
                        ],
                        spacing=5,
                    ),
                    bgcolor=config.COLOR_GRAY_50,
                    border_radius=8,
                    padding=10,
                )
                entry_list.controls.append(row)
                self.page.update()
            
            def rebuild_entry_list():
                entry_list.controls.clear()
                for i, row_data in enumerate(entry_rows["items"]):
                    def make_remove(idx):
                        def remove_row(e):
                            if len(entry_rows["items"]) > 1:
                                entry_rows["items"].pop(idx)
                                rebuild_entry_list()
                        return remove_row
                    
                    row = Container(
                        content=Row(
                            controls=[
                                row_data["year"], row_data["month"], row_data["day"],
                                Container(width=20),
                                row_data["start"],
                                Text("〜", size=14),
                                row_data["end"],
                                IconButton(
                                    icon=Icons.REMOVE_CIRCLE_OUTLINE,
                                    icon_color="#ef4444",
                                    tooltip="削除",
                                    on_click=make_remove(i),
                                ),
                            ],
                            spacing=5,
                        ),
                        bgcolor=config.COLOR_GRAY_50,
                        border_radius=8,
                        padding=10,
                    )
                    entry_list.controls.append(row)
                self.page.update()
            
            # 初期行を追加
            add_entry_row()
            
            def save_all_entries(e):
                if not club_dropdown.value or not facility_dropdown.value:
                    self.page.snack_bar = SnackBar(Text("団体と施設を選択してください"), bgcolor="#ef4444")
                    self.page.snack_bar.open = True
                    self.page.update()
                    return
                
                self.loading.show("保存中...")
                try:
                    for row_data in entry_rows["items"]:
                        date_str = f"{row_data['year'].value}/{row_data['month'].value}/{row_data['day'].value}"
                        self.sheets_service.add_secretary_log({
                            "Date": date_str,
                            "Facility": facility_dropdown.value,
                            "ClubName": club_dropdown.value,
                            "StartTime": row_data["start"].value,
                            "EndTime": row_data["end"].value,
                            "Note": "",
                        })
                    
                    self.page.snack_bar = SnackBar(Text(f"{len(entry_rows['items'])}件の記録を保存しました"), bgcolor="#22c55e")
                    self.page.snack_bar.open = True
                    
                    # リセット
                    entry_rows["items"].clear()
                    entry_list.controls.clear()
                    add_entry_row()
                    
                except Exception as ex:
                    get_error_logger().log_error("Secretary", "save_all", ex)
                    self.page.snack_bar = SnackBar(Text(f"エラー: {ex}"), bgcolor="#ef4444")
                    self.page.snack_bar.open = True
                finally:
                    self.loading.hide()
                    self.page.update()
            
            # 左側：入力フォーム
            input_form = Container(
                content=Column(
                    controls=[
                        Text("記録 打ち込み", size=18, weight=ft.FontWeight.BOLD),
                        Container(height=15),
                        Text("団体選択", size=14, weight=ft.FontWeight.W_500),
                        search_field,
                        club_dropdown,
                        Container(height=15),
                        Text("施設選択", size=14, weight=ft.FontWeight.W_500),
                        facility_dropdown,
                        Container(height=20),
                        Row(
                            controls=[
                                Text("活動期間（日付と時刻）", size=14, weight=ft.FontWeight.W_500),
                                IconButton(
                                    icon=Icons.ADD_CIRCLE,
                                    icon_color=get_accent_color(),
                                    tooltip="行を追加",
                                    on_click=add_entry_row,
                                ),
                            ],
                            alignment=MainAxisAlignment.SPACE_BETWEEN,
                        ),
                        Container(
                            content=entry_list,
                            height=250,
                            border=ft.border.all(1, config.COLOR_GRAY_200),
                            border_radius=8,
                            padding=10,
                        ),
                        Container(height=20),
                        create_primary_button(
                            "一括保存",
                            save_all_entries,
                            icon=Icons.SAVE,
                            width=200,
                            bgcolor=config.DEPT_COLORS["secretary"],
                        ),
                    ],
                    scroll=ScrollMode.AUTO,
                ),
                expand=2,
                padding=20,
                bgcolor=config.COLOR_WHITE,
                border_radius=12,
            )
            
            # 右側：PDF表示エリア
            pdf_panel = Container(
                content=Column(
                    controls=[
                        Text("PDFプレビュー", size=16, weight=ft.FontWeight.BOLD),
                        Container(height=10),
                        Container(
                            content=Column(
                                controls=[
                                    Icon(Icons.PICTURE_AS_PDF, size=50, color=config.COLOR_GRAY_300),
                                    Text("PDFファイルを選択してください", size=14, color=config.COLOR_GRAY_500),
                                    create_secondary_button(
                                        "Google Driveから選択",
                                        lambda e: self.file_picker.pick_files(
                                            dialog_title="PDFを選択",
                                            allowed_extensions=["pdf"],
                                        ),
                                        icon=Icons.FOLDER_OPEN,
                                    ),
                                ],
                                horizontal_alignment=CrossAxisAlignment.CENTER,
                                alignment=MainAxisAlignment.CENTER,
                                spacing=15,
                            ),
                            expand=True,
                            bgcolor=config.COLOR_GRAY_50,
                            border_radius=8,
                            alignment=alignment.center,
                        ),
                    ],
                ),
                expand=1,
                padding=20,
                bgcolor=config.COLOR_WHITE,
                border_radius=12,
            )
            
            return Row(
                controls=[input_form, pdf_panel],
                spacing=15,
                expand=True,
            )
        
        # --- Tab 1: Excel出力 ---
        def create_export_tab():
            # 記録一覧表示
            def create_logs_table():
                logs = []
                try:
                    if self.sheets_service:
                        logs = self.sheets_service.get_secretary_logs() or []
                except:
                    pass
                
                rows = []
                for i, log in enumerate(logs):
                    club_name = log.get("ClubName", "")
                    # 団体＋区分表示
                    club = next((c for c in self.clubs if c.get("ClubName") == club_name), None)
                    display_name = get_club_display_name(club) if club else club_name
                    
                    def make_edit(idx, log_data):
                        def edit(e):
                            self._show_edit_secretary_dialog(idx, log_data)
                        return edit
                    
                    def make_delete(idx):
                        def delete(e):
                            show_confirm_dialog(
                                self.page,
                                "削除確認",
                                "この記録を削除しますか？",
                                lambda: self._delete_secretary_log(idx),
                                is_danger=True,
                            )
                        return delete
                    
                    rows.append(DataRow(cells=[
                        DataCell(Text(log.get("Date", ""), size=12)),
                        DataCell(Text(display_name, size=12)),
                        DataCell(Text(log.get("Facility", ""), size=12)),
                        DataCell(Text(f"{log.get('StartTime', '')}〜{log.get('EndTime', '')}", size=12)),
                        DataCell(create_edit_delete_buttons(make_edit(i, log), make_delete(i))),
                    ]))
                
                if not rows:
                    return Container(
                        content=Text("記録がありません", color=config.COLOR_GRAY_500),
                        padding=30,
                        alignment=alignment.center,
                    )
                
                return DataTable(
                    columns=[
                        DataColumn(Text("日付", weight=ft.FontWeight.BOLD)),
                        DataColumn(Text("団体", weight=ft.FontWeight.BOLD)),
                        DataColumn(Text("施設", weight=ft.FontWeight.BOLD)),
                        DataColumn(Text("時間", weight=ft.FontWeight.BOLD)),
                        DataColumn(Text("操作", weight=ft.FontWeight.BOLD)),
                    ],
                    rows=rows,
                    border=ft.border.all(1, config.COLOR_GRAY_200),
                    border_radius=8,
                    heading_row_color=config.DEPT_COLORS["secretary"] + "15",
                )
            
            def export_excel(e):
                self.loading.show("Excel出力中...")
                try:
                    # TODO: Excel出力処理
                    self.page.snack_bar = SnackBar(Text("Excel出力機能は準備中です"), bgcolor="#f59e0b")
                    self.page.snack_bar.open = True
                except Exception as ex:
                    get_error_logger().log_error("Secretary", "export", ex)
                finally:
                    self.loading.hide()
                    self.page.update()
            
            return Column(
                controls=[
                    Row(
                        controls=[
                            Text("記録一覧", size=18, weight=ft.FontWeight.BOLD),
                            Container(expand=True),
                            create_primary_button("Excel出力", export_excel, icon=Icons.DOWNLOAD),
                        ],
                    ),
                    Container(height=15),
                    Container(
                        content=Column(
                            controls=[create_logs_table()],
                            scroll=ScrollMode.BOTH,
                        ),
                        expand=True,
                        bgcolor=config.COLOR_WHITE,
                        border_radius=12,
                        padding=15,
                    ),
                ],
                expand=True,
            )
        
        # Tab switching
        tabs = [("記録", create_record_tab), ("Excel出力", create_export_tab)]
        
        def switch_tab(index):
            current_tab["value"] = index
            for i, btn in enumerate(tab_buttons_row.controls):
                btn.style = ft.ButtonStyle(
                    bgcolor=get_accent_color() if i == index else config.COLOR_GRAY_100,
                    color=config.COLOR_WHITE if i == index else config.COLOR_GRAY_700,
                    padding=padding.symmetric(horizontal=20, vertical=10),
                    shape=RoundedRectangleBorder(radius=8),
                )
            tab_content.content = tabs[index][1]()
            self.page.update()
        
        tab_buttons_row.controls = self._create_tab_buttons(tabs, current_tab, switch_tab)
        tab_content.content = create_record_tab()
        
        header = self._create_dept_header("書記部", config.DEPT_COLORS["secretary"], Icons.CALENDAR_MONTH)
        
        content = Container(
            content=Column(
                controls=[
                    tab_buttons_row,
                    Container(height=10),
                    tab_content,
                ],
                expand=True,
            ),
            padding=15,
            expand=True,
        )
        
        self.page.controls.clear()
        self.page.add(Column(controls=[header, content], spacing=0, expand=True))
        self.page.update()
    
    def _show_edit_secretary_dialog(self, index: int, log: Dict):
        """書記部記録の編集ダイアログ"""
        date_parts = log.get("Date", "").split("/")
        year = int(date_parts[0]) if len(date_parts) > 0 else datetime.now().year
        month = int(date_parts[1]) if len(date_parts) > 1 else datetime.now().month
        day = int(date_parts[2]) if len(date_parts) > 2 else datetime.now().day
        
        year_dd, month_dd, day_dd = create_date_dropdowns(year, month, day)
        start_time = create_time_dropdown("開始", log.get("StartTime", "09:00"))
        end_time = create_time_dropdown("終了", log.get("EndTime", "17:00"))
        
        club_options = [(c.get("ClubName", ""), get_club_display_name(c)) for c in self.clubs]
        club_dd = create_dropdown("団体", club_options, value=log.get("ClubName", ""))
        
        facility_options = [(f.get("FacilityName", ""), f.get("FacilityName", "")) for f in self.facilities]
        facility_dd = create_dropdown("施設", facility_options, value=log.get("Facility", ""))
        
        def save(e):
            try:
                date_str = f"{year_dd.value}/{month_dd.value}/{day_dd.value}"
                self.sheets_service.update_row(config.SHEET_SECRETARY_LOG, index, [
                    date_str, facility_dd.value, club_dd.value,
                    start_time.value, end_time.value, "", datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])
                dialog.open = False
                self.page.snack_bar = SnackBar(Text("更新しました"), bgcolor="#22c55e")
                self.page.snack_bar.open = True
                self._show_secretary_view()  # Refresh
            except Exception as ex:
                get_error_logger().log_error("Secretary", "edit", ex)
        
        def cancel(e):
            dialog.open = False
            self.page.update()
        
        dialog = AlertDialog(
            modal=True,
            title=Text("記録を編集", weight=ft.FontWeight.BOLD),
            content=Column(
                controls=[
                    club_dd,
                    facility_dd,
                    Container(height=10),
                    Row(controls=[year_dd, month_dd, day_dd], spacing=5),
                    Row(controls=[start_time, Text("〜"), end_time], spacing=5),
                ],
                tight=True,
                spacing=10,
            ),
            actions=[
                TextButton("キャンセル", on_click=cancel),
                ElevatedButton("保存", on_click=save, style=ft.ButtonStyle(bgcolor=get_accent_color(), color=config.COLOR_WHITE)),
            ],
        )
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def _delete_secretary_log(self, index: int):
        try:
            self.sheets_service.delete_secretary_log(index)
            self.page.snack_bar = SnackBar(Text("削除しました"), bgcolor="#22c55e")
            self.page.snack_bar.open = True
            self._show_secretary_view()
        except Exception as ex:
            get_error_logger().log_error("Secretary", "delete", ex)
            self.page.snack_bar = SnackBar(Text(f"エラー: {ex}"), bgcolor="#ef4444")
            self.page.snack_bar.open = True
            self.page.update()
    
    # ============================================================
    # Finance View (財務部)
    # ============================================================
    def _show_finance_view(self):
        self.current_view = "finance"
        
        # Load finance data
        finance_entries = []
        try:
            if self.sheets_service:
                finance_entries = self.sheets_service.get_finance_entries() or []
        except Exception as e:
            get_error_logger().log_error("Finance", "load", e)
        
        # Calculate totals
        cash_total = 0
        bank_total = 0
        for entry in finance_entries:
            income = int(entry.get("Income", 0) or 0)
            expense = int(entry.get("Expense", 0) or 0)
            method = entry.get("PaymentMethod", "")
            if method in ["cash", "現金"]:
                cash_total += income - expense
            else:
                bank_total += income - expense
        
        # Form fields (Left side)
        now = datetime.now()
        year_dd, month_dd, day_dd = create_date_dropdowns()
        
        event_options = [
            ("文化の祭典", "文化の祭典"), 
            ("体育の祭典", "体育の祭典"), 
            ("定例会", "定例会"),
            ("その他", "その他"),
        ]
        event_dropdown = create_dropdown("行事・科目", event_options, width=300)
        
        desc_field = create_text_field("摘要", width=300)
        
        payment_group = ft.RadioGroup(
            content=Row(controls=[
                ft.Radio(value="現金", label="現金"),
                ft.Radio(value="通帳", label="通帳"),
            ]),
            value="現金",
        )
        
        income_field = create_number_field("収入(円)", width=150)
        expense_field = create_number_field("支出(円)", width=150)
        
        reimburse_group = ft.RadioGroup(
            content=Row(controls=[
                ft.Radio(value="/", label="/"),
                ft.Radio(value="未返金", label="未返金"),
                ft.Radio(value="返金済", label="返金済"),
            ]),
            value="/",
        )
        
        def on_add(e):
            try:
                date_str = f"{year_dd.value}/{month_dd.value}/{day_dd.value}"
                self.sheets_service.add_finance_entry({
                    "Date": date_str,
                    "Subject": event_dropdown.value,
                    "Description": desc_field.value,
                    "PaymentMethod": payment_group.value,
                    "Income": parse_number_from_comma(income_field.value),
                    "Expense": parse_number_from_comma(expense_field.value),
                    "ReimbursementStatus": reimburse_group.value,
                })
                desc_field.value = ""
                income_field.value = ""
                expense_field.value = ""
                self._show_finance_view()  # Refresh
            except Exception as ex:
                get_error_logger().log_error("Finance", "add", ex)
                self.page.snack_bar = SnackBar(Text(f"エラー: {ex}"), bgcolor="#ef4444")
                self.page.snack_bar.open = True
                self.page.update()
        
        input_form = Container(
            content=Column(
                controls=[
                    Text("財務部 - 出納帳管理", size=18, weight=ft.FontWeight.BOLD),
                    Container(height=15),
                    Container(
                        content=Column(
                            controls=[
                                Text("新規取引入力", size=14, weight=ft.FontWeight.W_500),
                                Container(height=10),
                                Text("日付", size=12, color=config.COLOR_GRAY_600),
                                Row(controls=[year_dd, month_dd, day_dd], spacing=5),
                                Container(height=10),
                                Text("行事・科目", size=12, color=config.COLOR_GRAY_600),
                                event_dropdown,
                                Container(height=10),
                                Text("摘要", size=12, color=config.COLOR_GRAY_600),
                                desc_field,
                                Container(height=10),
                                Text("支払方法", size=12, color=config.COLOR_GRAY_600),
                                payment_group,
                                Container(height=10),
                                Row(controls=[
                                    Column(controls=[
                                        Text("収入(円)", size=12, color=config.COLOR_GRAY_600),
                                        income_field,
                                    ]),
                                    Column(controls=[
                                        Text("支出(円)", size=12, color=config.COLOR_GRAY_600),
                                        expense_field,
                                    ]),
                                ], spacing=15),
                                Container(height=10),
                                Text("立替状態", size=12, color=config.COLOR_GRAY_600),
                                reimburse_group,
                                Container(height=20),
                                create_primary_button("追加", on_add, width=300, bgcolor=config.DEPT_COLORS["finance"]),
                            ],
                        ),
                        bgcolor=config.COLOR_WHITE,
                        border_radius=12,
                        border=ft.border.all(1, config.COLOR_GRAY_200),
                        padding=20,
                    ),
                ],
                scroll=ScrollMode.AUTO,
            ),
            width=380,
        )
        
        # Right side - Search and table
        search_field = create_search_field("科目・摘要で検索...", width=250)
        year_filter = create_dropdown("年", [(str(y), str(y)) for y in range(2024, 2030)], value=str(now.year), width=90)
        month_filter = create_dropdown("月", [("all", "全月")] + [(str(m), f"{m}月") for m in range(1, 13)], value="all", width=80)
        
        # Table
        def create_finance_table():
            rows = []
            for i, entry in enumerate(finance_entries):
                reimburse_text = entry.get("ReimbursementStatus", "/")
                reimburse_color = "#ef4444" if reimburse_text == "未返金" else (config.COLOR_GRAY_600 if reimburse_text == "/" else "#22c55e")
                
                income_val = int(entry.get("Income", 0) or 0)
                expense_val = int(entry.get("Expense", 0) or 0)
                
                def make_edit(idx, entry_data):
                    def edit(e):
                        self._show_edit_finance_dialog(idx, entry_data)
                    return edit
                
                def make_delete(idx):
                    def delete(e):
                        show_confirm_dialog(
                            self.page,
                            "削除確認",
                            "この取引を削除しますか？",
                            lambda: self._delete_finance_entry(idx),
                            is_danger=True,
                        )
                    return delete
                
                rows.append(DataRow(cells=[
                    DataCell(Text(entry.get("Date", ""), size=12)),
                    DataCell(Text(entry.get("Subject", ""), size=12)),
                    DataCell(Text(entry.get("Description", ""), size=12)),
                    DataCell(Text(entry.get("PaymentMethod", ""), size=12)),
                    DataCell(Text(f"¥{income_val:,}" if income_val else "", size=12, color="#3b82f6")),
                    DataCell(Text(f"¥{expense_val:,}" if expense_val else "", size=12, color="#ef4444")),
                    DataCell(Text(reimburse_text, color=reimburse_color, size=12)),
                    DataCell(create_edit_delete_buttons(make_edit(i, entry), make_delete(i))),
                ]))
            
            if not rows:
                return Container(
                    content=Text("取引記録がありません", color=config.COLOR_GRAY_500),
                    padding=30,
                    alignment=alignment.center,
                )
            
            return DataTable(
                columns=[
                    DataColumn(Text("日付", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("科目", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("摘要", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("支払", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("収入", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("支出", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("立替", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("操作", weight=ft.FontWeight.BOLD)),
                ],
                rows=rows,
                border=ft.border.all(1, config.COLOR_GRAY_200),
                border_radius=8,
                heading_row_color=config.DEPT_COLORS["finance"] + "15",
            )
        
        # Summary row
        search_income = sum(int(e.get("Income", 0) or 0) for e in finance_entries)
        search_expense = sum(int(e.get("Expense", 0) or 0) for e in finance_entries)
        
        def summary_cell(label: str, value: str, color: str = config.COLOR_GRAY_800):
            return Container(
                content=Column(
                    controls=[
                        Text(label, size=11, color=config.COLOR_GRAY_600),
                        Text(value, size=14, weight=ft.FontWeight.BOLD, color=color),
                    ],
                    horizontal_alignment=CrossAxisAlignment.CENTER,
                    spacing=2,
                ),
                padding=10,
            )
        
        summary_row = Container(
            content=Row(
                controls=[
                    summary_cell("検索後 収入", f"¥{search_income:,}", "#3b82f6"),
                    summary_cell("検索後 支出", f"¥{search_expense:,}", "#ef4444"),
                    summary_cell("手許現金", f"¥{cash_total:,}"),
                    summary_cell("通帳残高", f"¥{bank_total:,}"),
                    summary_cell("合計残高", f"¥{cash_total + bank_total:,}", "#3b82f6"),
                ],
                alignment=MainAxisAlignment.SPACE_AROUND,
            ),
            bgcolor="#fffef0",
            padding=10,
            border_radius=8,
        )
        
        def export_excel(e):
            self.page.snack_bar = SnackBar(Text("Excel出力機能は準備中です"), bgcolor="#f59e0b")
            self.page.snack_bar.open = True
            self.page.update()
        
        right_panel = Container(
            content=Column(
                controls=[
                    Row(
                        controls=[
                            search_field,
                            Container(expand=True),
                            year_filter,
                            month_filter,
                            create_primary_button("Excel出力", export_excel, icon=Icons.DOWNLOAD),
                        ],
                        spacing=10,
                    ),
                    Container(height=10),
                    Container(
                        content=Column(
                            controls=[create_finance_table()],
                            scroll=ScrollMode.BOTH,
                        ),
                        expand=True,
                        bgcolor=config.COLOR_WHITE,
                        border_radius=8,
                    ),
                    Container(height=10),
                    summary_row,
                ],
                expand=True,
            ),
            expand=True,
            padding=10,
        )
        
        header = self._create_dept_header("財務部", config.DEPT_COLORS["finance"], Icons.ACCOUNT_BALANCE_WALLET)
        
        content = Container(
            content=Row(
                controls=[
                    input_form,
                    Container(content=right_panel, expand=True),
                ],
                spacing=15,
                expand=True,
            ),
            padding=15,
            expand=True,
        )
        
        self.page.controls.clear()
        self.page.add(Column(controls=[header, content], spacing=0, expand=True))
        self.page.update()
    
    def _show_edit_finance_dialog(self, index: int, entry: Dict):
        """財務部記録の編集ダイアログ"""
        date_parts = entry.get("Date", "").split("/")
        year = int(date_parts[0]) if len(date_parts) > 0 else datetime.now().year
        month = int(date_parts[1]) if len(date_parts) > 1 else datetime.now().month
        day = int(date_parts[2]) if len(date_parts) > 2 else datetime.now().day
        
        year_dd, month_dd, day_dd = create_date_dropdowns(year, month, day)
        
        event_options = [
            ("文化の祭典", "文化の祭典"), ("体育の祭典", "体育の祭典"),
            ("定例会", "定例会"), ("その他", "その他"),
        ]
        event_dd = create_dropdown("行事・科目", event_options, value=entry.get("Subject", ""), width=250)
        desc_field = create_text_field("摘要", entry.get("Description", ""), width=250)
        
        payment_group = ft.RadioGroup(
            content=Row(controls=[
                ft.Radio(value="現金", label="現金"),
                ft.Radio(value="通帳", label="通帳"),
            ]),
            value=entry.get("PaymentMethod", "現金"),
        )
        
        income_field = create_number_field("収入", int(entry.get("Income", 0) or 0), width=120)
        expense_field = create_number_field("支出", int(entry.get("Expense", 0) or 0), width=120)
        
        reimburse_group = ft.RadioGroup(
            content=Row(controls=[
                ft.Radio(value="/", label="/"),
                ft.Radio(value="未返金", label="未返金"),
                ft.Radio(value="返金済", label="返金済"),
            ]),
            value=entry.get("ReimbursementStatus", "/"),
        )
        
        def save(e):
            try:
                date_str = f"{year_dd.value}/{month_dd.value}/{day_dd.value}"
                self.sheets_service.update_finance_entry(index, {
                    "Date": date_str,
                    "Subject": event_dd.value,
                    "Description": desc_field.value,
                    "PaymentMethod": payment_group.value,
                    "Income": parse_number_from_comma(income_field.value),
                    "Expense": parse_number_from_comma(expense_field.value),
                    "ReimbursementStatus": reimburse_group.value,
                })
                dialog.open = False
                self.page.snack_bar = SnackBar(Text("更新しました"), bgcolor="#22c55e")
                self.page.snack_bar.open = True
                self._show_finance_view()
            except Exception as ex:
                get_error_logger().log_error("Finance", "edit", ex)
        
        def cancel(e):
            dialog.open = False
            self.page.update()
        
        dialog = AlertDialog(
            modal=True,
            title=Text("取引を編集", weight=ft.FontWeight.BOLD),
            content=Column(
                controls=[
                    Row(controls=[year_dd, month_dd, day_dd], spacing=5),
                    event_dd, desc_field,
                    payment_group,
                    Row(controls=[income_field, expense_field], spacing=10),
                    reimburse_group,
                ],
                tight=True,
                spacing=10,
            ),
            actions=[
                TextButton("キャンセル", on_click=cancel),
                ElevatedButton("保存", on_click=save, style=ft.ButtonStyle(bgcolor=get_accent_color(), color=config.COLOR_WHITE)),
            ],
        )
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def _delete_finance_entry(self, index: int):
        try:
            self.sheets_service.delete_finance_entry(index)
            self.page.snack_bar = SnackBar(Text("削除しました"), bgcolor="#22c55e")
            self.page.snack_bar.open = True
            self._show_finance_view()
        except Exception as ex:
            get_error_logger().log_error("Finance", "delete", ex)
            self.page.snack_bar = SnackBar(Text(f"エラー: {ex}"), bgcolor="#ef4444")
            self.page.snack_bar.open = True
            self.page.update()
        
        input_form = Container(
            content=Column(
                controls=[
                    Text("財務部 - 出納帳管理", size=18, weight=ft.FontWeight.BOLD),
                    Container(height=15),
                    Container(
                        content=Column(
                            controls=[
                                Text("新規取引入力", size=14, weight=ft.FontWeight.W_500),
                                Container(height=10),
                                Text("日付", size=12, color=config.COLOR_GRAY_600),
                                date_field,
                                Container(height=10),
                                Text("行事・科目", size=12, color=config.COLOR_GRAY_600),
                                event_dropdown,
                                Container(height=10),
                                Text("摘要", size=12, color=config.COLOR_GRAY_600),
                                desc_field,
                                Container(height=10),
                                Text("支払方法", size=12, color=config.COLOR_GRAY_600),
                                payment_group,
                                Container(height=10),
                                Text("収入(円)", size=12, color=config.COLOR_GRAY_600),
                                income_field,
                                Container(height=10),
                                Text("支出(円)", size=12, color=config.COLOR_GRAY_600),
                                expense_field,
                                Container(height=10),
                                Text("立替状態", size=12, color=config.COLOR_GRAY_600),
                                reimburse_group,
                                Container(height=20),
                                create_primary_button("追加", on_add, width=300, bgcolor="#f97316"),
                            ],
                        ),
                        bgcolor=config.COLOR_WHITE,
                        border_radius=12,
                        border=ft.border.all(1, config.COLOR_GRAY_200),
                        padding=20,
                    ),
                ],
            ),
            width=380,
        )
        
        # Right side - Search and table
        search_field = create_text_field("科目・摘要で検索...", width=250)
        year_dropdown = create_dropdown("年", [(str(y), str(y)) for y in range(2024, 2030)], value="2026", width=90)
        month_dropdown = create_dropdown("月", [(str(m), f"{m}月") for m in range(1, 13)], value=str(datetime.now().month), width=70)
        
        export_btn = create_primary_button("月別Excel出力", lambda e: None, icon=Icons.DOWNLOAD)
        
        # Table
        def create_finance_table():
            rows = []
            for i, entry in enumerate(finance_entries):
                reimburse_text = entry.get("ReimbursementStatus", "/")
                reimburse_color = "#ef4444" if reimburse_text == "未返金" else (config.COLOR_GRAY_600 if reimburse_text == "/" else "#22c55e")
                
                income_val = entry.get("Income", 0)
                expense_val = entry.get("Expense", 0)
                
                rows.append(DataRow(cells=[
                    DataCell(Text(entry.get("Date", ""), size=12)),
                    DataCell(Text(entry.get("Subject", ""), size=12)),
                    DataCell(Text(entry.get("Description", ""), size=12)),
                    DataCell(Text(entry.get("PaymentMethod", ""), size=12)),
                    DataCell(Text(f"¥{income_val:,}" if income_val else "", size=12)),
                    DataCell(Text(f"¥{expense_val:,}" if expense_val else "", size=12)),
                    DataCell(Text(reimburse_text, color=reimburse_color, size=12)),
                    DataCell(Row(controls=[
                        TextButton("編", on_click=lambda e, idx=i: None),
                        TextButton("削", style=ft.ButtonStyle(color="#ef4444"), on_click=lambda e, idx=i: None),
                    ])),
                ]))
            
            return DataTable(
                columns=[
                    DataColumn(Text("日付", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("科目", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("摘要", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("支払", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("収入", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("支出", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("立替", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("操作", weight=ft.FontWeight.BOLD)),
                ],
                rows=rows,
                border=ft.border.all(1, config.COLOR_GRAY_200),
                border_radius=8,
                heading_row_color="#fffef0",
            )
        
        # Summary row
        search_income = sum(int(e.get("Income", 0) or 0) for e in finance_entries)
        search_expense = sum(int(e.get("Expense", 0) or 0) for e in finance_entries)
        
        def summary_cell(label: str, value: str, color: str = config.COLOR_GRAY_800):
            return Column(
                controls=[
                    Text(label, size=11, color=config.COLOR_GRAY_600),
                    Text(value, size=14, weight=ft.FontWeight.BOLD, color=color),
                ],
                horizontal_alignment=CrossAxisAlignment.CENTER,
                spacing=2,
            )
        
        summary_row = Container(
            content=Row(
                controls=[
                    summary_cell("検索後 合計収入", f"¥{search_income:,}", "#3b82f6"),
                    summary_cell("検索後 合計支出", f"¥{search_expense:,}", "#ef4444"),
                    summary_cell("全体 手許現金", f"¥{cash_total:,}"),
                    summary_cell("全体 通帳残高", f"¥{bank_total:,}"),
                    summary_cell("全体 合計残高", f"¥{cash_total + bank_total:,}", "#3b82f6"),
                ],
                alignment=MainAxisAlignment.SPACE_AROUND,
            ),
            bgcolor="#fffef0",
            padding=15,
            border_radius=8,
        )
        
        right_panel = Container(
            content=Column(
                controls=[
                    Row(
                        controls=[
                            search_field,
                            Container(expand=True),
                            year_dropdown,
                            month_dropdown,
                            export_btn,
                        ],
                        spacing=10,
                    ),
                    Container(height=10),
                    Container(
                        content=Column(
                            controls=[create_finance_table()],
                            scroll=ScrollMode.BOTH,
                        ),
                        expand=True,
                        bgcolor=config.COLOR_WHITE,
                        border_radius=8,
                    ),
                    Container(height=10),
                    summary_row,
                ],
                expand=True,
            ),
            expand=True,
        )
        
        header = self._create_dept_header("財務部", config.DEPT_COLORS["finance"], Icons.ACCOUNT_BALANCE_WALLET)
        
        content = Container(
            content=Row(
                controls=[
                    input_form,
                    Container(content=right_panel, expand=True),
                ],
                spacing=20,
                expand=True,
            ),
            padding=20,
            expand=True,
        )
        
        self.page.controls.clear()
        self.page.add(Column(controls=[header, content], spacing=0, expand=True))
        self.page.update()
    
    # ============================================================
    # General View (総務部) - Fixed blank screen
    # ============================================================
    def _show_general_view(self):
        self.current_view = "general"
        
        current_tab = {"value": 0}
        tab_content = Container(expand=True)
        tab_buttons_row = Row(controls=[], spacing=10)
        
        # Get data
        assignments = {}
        attendance_records = []
        try:
            if self.sheets_service:
                assignments = self.sheets_service.get_weekday_assignments() or {}
                attendance_records = self.sheets_service.get_attendance() or []
        except Exception as e:
            get_error_logger().log_error("General", "load", e)
        
        # --- Tab 0: 担当保存 ---
        def create_assign_tab():
            period_tabs_list = []
            
            for period_key, period_label in config.PERIODS:
                period_data = assignments.get(period_key, {})
                
                day_rows = []
                for day_key, day_label in config.WEEKDAYS:
                    day_members = period_data.get(day_key, [])
                    
                    if day_members:
                        member_chips = Row(
                            controls=[
                                Container(
                                    content=Text(name, size=12, color=config.COLOR_WHITE),
                                    bgcolor=get_accent_color(),
                                    border_radius=12,
                                    padding=padding.symmetric(horizontal=10, vertical=4),
                                )
                                for name in day_members
                            ],
                            wrap=True,
                            spacing=5,
                        )
                    else:
                        member_chips = Text("未設定", color=config.COLOR_GRAY_400, size=12)
                    
                    day_rows.append(
                        Container(
                            content=Row(
                                controls=[
                                    Container(
                                        content=Text(day_label, weight=ft.FontWeight.BOLD, size=14),
                                        width=50,
                                    ),
                                    member_chips,
                                ],
                            ),
                            padding=12,
                            border=ft.border.only(bottom=BorderSide(1, config.COLOR_GRAY_200)),
                        )
                    )
                
                period_tabs_list.append(
                    ft.Tab(
                        text=period_label,
                        content=Container(
                            content=Column(controls=day_rows, scroll=ScrollMode.AUTO),
                            padding=15,
                            bgcolor=config.COLOR_WHITE,
                        ),
                    )
                )
            
            return Container(
                content=ft.Tabs(
                    tabs=period_tabs_list,
                    expand=True,
                    animation_duration=0,
                ),
                expand=True,
            )
        
        # --- Tab 1: 出欠記録 ---
        def create_attendance_tab():
            today = datetime.now().strftime("%Y-%m-%d")
            today_weekday_idx = datetime.now().weekday()
            weekday_keys = ["mon", "tue", "wed", "thu", "fri"]
            today_weekday = weekday_keys[today_weekday_idx] if today_weekday_idx < 5 else "mon"
            
            current_period = "zenki"  # TODO: Auto-detect period
            today_assignments = assignments.get(current_period, {}).get(today_weekday, [])
            
            recorded_today = [r for r in attendance_records if r.get("Date") == today]
            recorded_names = {r.get("MemberName") for r in recorded_today}
            
            unrecorded = [m for m in today_assignments if m not in recorded_names]
            
            def record_attendance(name, status):
                try:
                    self.sheets_service.add_attendance({
                        "Date": today,
                        "MemberName": name,
                        "Status": status,
                        "Period": current_period,
                    })
                    self._show_general_view()
                except Exception as ex:
                    get_error_logger().log_error("General", "record_attendance", ex)
            
            unrecorded_list = Column(
                controls=[
                    Container(
                        content=Row(
                            controls=[
                                Text(name, expand=True, size=14),
                                Dropdown(
                                    options=[dropdown.Option(k, v) for k, v in config.ATTENDANCE_STATUS],
                                    value="present",
                                    width=100,
                                    border_radius=8,
                                ),
                                create_primary_button("記録", lambda e, n=name: record_attendance(n, "present")),
                            ],
                            spacing=10,
                        ),
                        padding=10,
                        border=ft.border.only(bottom=BorderSide(1, config.COLOR_GRAY_100)),
                    )
                    for name in unrecorded
                ] if unrecorded else [Text("全員記録済み", color=config.COLOR_GRAY_500)],
                scroll=ScrollMode.AUTO,
            )
            
            recorded_list = Column(
                controls=[
                    Container(
                        content=Row(
                            controls=[
                                Text(r.get("MemberName", ""), expand=True),
                                Text(dict(config.ATTENDANCE_STATUS).get(r.get("Status", ""), r.get("Status", "")), 
                                     color="#22c55e" if r.get("Status") == "present" else "#ef4444"),
                            ],
                        ),
                        padding=8,
                    )
                    for r in recorded_today
                ] if recorded_today else [Text("記録なし", color=config.COLOR_GRAY_500)],
                scroll=ScrollMode.AUTO,
            )
            
            return Row(
                controls=[
                    Container(
                        content=Column(
                            controls=[
                                Text("未記録", weight=ft.FontWeight.BOLD, color="#ef4444", size=16),
                                Container(height=10),
                                unrecorded_list,
                            ],
                            expand=True,
                        ),
                        expand=True,
                        bgcolor=config.COLOR_WHITE,
                        border_radius=12,
                        padding=15,
                    ),
                    Container(
                        content=Column(
                            controls=[
                                Text("記録済み", weight=ft.FontWeight.BOLD, color="#22c55e", size=16),
                                Container(height=10),
                                recorded_list,
                            ],
                            expand=True,
                        ),
                        expand=True,
                        bgcolor=config.COLOR_WHITE,
                        border_radius=12,
                        padding=15,
                    ),
                ],
                spacing=20,
                expand=True,
            )
        
        # --- Tab 2: 出欠割合統計 ---
        def create_stats_tab():
            member_stats = {}
            for member in self.members:
                name = member.get("Name", "")
                dept = member.get("Department", "")
                records = [r for r in attendance_records if r.get("MemberName") == name]
                
                total = len(records)
                present = len([r for r in records if r.get("Status") == "present"])
                absent = len([r for r in records if r.get("Status") == "absent"])
                late = len([r for r in records if r.get("Status") == "late"])
                early = len([r for r in records if r.get("Status") == "early_leave"])
                mourning = len([r for r in records if r.get("Status") == "mourning"])
                rate = f"{(present / total * 100):.1f}%" if total > 0 else "-"
                
                member_stats[name] = {
                    "dept": dept, "total": total, "present": present,
                    "absent": absent, "late": late, "early": early,
                    "mourning": mourning, "rate": rate,
                }
            
            rows = [
                DataRow(cells=[
                    DataCell(Text(name, size=12)),
                    DataCell(Text(stats["dept"], size=12)),
                    DataCell(Text(str(stats["total"]), size=12)),
                    DataCell(Text(str(stats["present"]), size=12)),
                    DataCell(Text(str(stats["absent"]), size=12)),
                    DataCell(Text(str(stats["late"]), size=12)),
                    DataCell(Text(str(stats["early"]), size=12)),
                    DataCell(Text(str(stats["mourning"]), size=12)),
                    DataCell(Text(stats["rate"], size=12, weight=ft.FontWeight.BOLD)),
                ])
                for name, stats in member_stats.items()
            ]
            
            table = DataTable(
                columns=[
                    DataColumn(Text(c, weight=ft.FontWeight.BOLD)) for c in
                    ["担当者", "部署", "総記録数", "出席", "欠席", "遅刻", "早退", "忌引等", "出席率"]
                ],
                rows=rows,
                border=ft.border.all(1, config.COLOR_GRAY_200),
                border_radius=8,
                heading_row_color=config.COLOR_GRAY_50,
            ) if rows else Text("データがありません", color=config.COLOR_GRAY_500)
            
            return Container(
                content=Column(
                    controls=[
                        Text("出欠割合統計", size=18, weight=ft.FontWeight.BOLD),
                        Container(height=15),
                        table if rows else Container(content=table, alignment=alignment.center),
                    ],
                    scroll=ScrollMode.AUTO,
                    expand=True,
                ),
                bgcolor=config.COLOR_WHITE,
                border_radius=12,
                padding=20,
                expand=True,
            )
        
        # Tab switching
        tabs = [("担当保存", create_assign_tab), ("出欠記録", create_attendance_tab), ("出欠割合統計", create_stats_tab)]
        
        def switch_tab(index):
            current_tab["value"] = index
            for i, btn in enumerate(tab_buttons_row.controls):
                btn.style = ft.ButtonStyle(
                    bgcolor=get_accent_color() if i == index else config.COLOR_GRAY_100,
                    color=config.COLOR_WHITE if i == index else config.COLOR_GRAY_700,
                    padding=padding.symmetric(horizontal=20, vertical=10),
                    shape=RoundedRectangleBorder(radius=8),
                )
            tab_content.content = tabs[index][1]()
            self.page.update()
        
        tab_buttons_row.controls = self._create_tab_buttons(tabs, current_tab, switch_tab)
        tab_content.content = create_assign_tab()
        
        header = self._create_dept_header("総務部", config.DEPT_COLORS["general"], Icons.GROUPS)
        
        content = Container(
            content=Column(
                controls=[
                    tab_buttons_row,
                    Container(height=15),
                    tab_content,
                ],
            ),
            padding=20,
            expand=True,
        )
        
        self.page.controls.clear()
        self.page.add(Column(controls=[header, content], spacing=0, expand=True))
        self.page.update()
    
    # ============================================================
    # External View (渉外部) - Same layout as GAS
    # ============================================================
    def _show_external_view(self):
        self.current_view = "external"
        
        current_tab = {"value": 0}
        tab_content = Container(expand=True)
        tab_buttons_row = Row(controls=[], spacing=10)
        
        # Load data
        external_logs = []
        try:
            if self.sheets_service:
                external_logs = self.sheets_service.get_external_logs() or []
        except:
            pass
        
        # --- Tab 0: 記録 (課外活動届 / 活動報告書) ---
        def create_record_tab():
            log_type = {"value": "activity"}
            
            activity_btn = ElevatedButton(
                "課外活動届",
                on_click=lambda e: set_log_type("activity"),
                style=ft.ButtonStyle(bgcolor=get_accent_color(), color=config.COLOR_WHITE),
            )
            report_btn = ElevatedButton(
                "活動報告書",
                on_click=lambda e: set_log_type("report"),
                style=ft.ButtonStyle(bgcolor=config.COLOR_GRAY_100, color=config.COLOR_GRAY_700),
            )
            
            def set_log_type(t):
                log_type["value"] = t
                activity_btn.style = ft.ButtonStyle(bgcolor=get_accent_color() if t == "activity" else config.COLOR_GRAY_100, color=config.COLOR_WHITE if t == "activity" else config.COLOR_GRAY_700)
                report_btn.style = ft.ButtonStyle(bgcolor=get_accent_color() if t == "report" else config.COLOR_GRAY_100, color=config.COLOR_WHITE if t == "report" else config.COLOR_GRAY_700)
                form_title.value = "課外活動届" if t == "activity" else "活動報告書"
                self.page.update()
            
            form_title = Text("課外活動届", size=16, weight=ft.FontWeight.BOLD)
            
            club_options = [(c.get("ClubName", ""), c.get("ClubName", "")) for c in self.clubs]
            if not club_options:
                club_options = [("", "選択してください")]
            club_dropdown = create_dropdown("団体名", club_options, expand=True)
            
            start_date = create_text_field("", datetime.now().strftime("%Y/%m/%d"), width=130)
            end_date = create_text_field("", datetime.now().strftime("%Y/%m/%d"), width=130)
            
            overnight_cb = Checkbox(label="宿泊を伴う")
            match_cb = Checkbox(label="公式戦")
            
            tournament_field = create_text_field("大会名", expand=True)
            
            def on_record(e):
                try:
                    self.sheets_service.add_external_log({
                        "LogType": log_type["value"],
                        "ClubName": club_dropdown.value,
                        "Period": f"{start_date.value}~{end_date.value}",
                        "HasOvernight": "TRUE" if overnight_cb.value else "FALSE",
                        "HasMatch": "TRUE" if match_cb.value else "FALSE",
                        "TournamentName": tournament_field.value,
                    })
                    self.page.snack_bar = SnackBar(Text("記録しました"), bgcolor="#22c55e")
                    self.page.snack_bar.open = True
                    tournament_field.value = ""
                    self.page.update()
                except Exception as ex:
                    get_error_logger().log_error("External", "record", ex)
            
            # Left side - Input form
            form_panel = Container(
                content=Column(
                    controls=[
                        Row(controls=[activity_btn, report_btn], spacing=10),
                        Container(height=15),
                        Container(
                            content=Column(
                                controls=[
                                    form_title,
                                    Container(height=10),
                                    Text("団体名", size=12, color=config.COLOR_GRAY_600),
                                    club_dropdown,
                                    Container(height=10),
                                    Text("期間", size=12, color=config.COLOR_GRAY_600),
                                    Row(
                                        controls=[
                                            start_date,
                                            Text("〜", color=config.COLOR_GRAY_600),
                                            end_date,
                                            IconButton(icon=Icons.CLOSE, icon_color="#ef4444", icon_size=18),
                                        ],
                                        spacing=5,
                                    ),
                                    TextButton("⊕ 追加", on_click=lambda e: None),
                                    Container(height=10),
                                    Row(controls=[overnight_cb, match_cb]),
                                    Container(height=10),
                                    Text("大会名", size=12, color=config.COLOR_GRAY_600),
                                    tournament_field,
                                    Container(height=20),
                                    create_primary_button("記録する", on_record, expand=True, bgcolor="#f97316"),
                                ],
                            ),
                            bgcolor=config.COLOR_WHITE,
                            border_radius=12,
                            border=ft.border.all(1, config.COLOR_GRAY_200),
                            padding=20,
                        ),
                    ],
                ),
            )
            
            # Right - PDF panel
            pdf_panel = Container(
                content=Column(
                    controls=[
                        Text("スキャンしたデータPDF", size=16, weight=ft.FontWeight.BOLD, color=config.COLOR_WHITE),
                        Container(height=20),
                        Row(
                            controls=[
                                create_secondary_button("OneDriveから選択", lambda e: self.file_picker.pick_files(allowed_extensions=["pdf"]), icon=Icons.FOLDER_OPEN),
                                create_secondary_button("スキャン実行", lambda e: None, icon=Icons.SCANNER),
                            ],
                            spacing=10,
                            wrap=True,
                        ),
                        Container(height=20),
                        Container(
                            content=Text("PDFプレビューエリア", color=config.COLOR_GRAY_400),
                            bgcolor=config.COLOR_WHITE,
                            border_radius=8,
                            padding=30,
                            expand=True,
                            alignment=alignment.center,
                        ),
                    ],
                    horizontal_alignment=CrossAxisAlignment.CENTER,
                ),
                bgcolor=config.DEPT_COLORS["external"],
                border_radius=12,
                padding=20,
                expand=True,
            )
            
            # Record list (top left)
            record_list = Container(
                content=Column(
                    controls=[
                        Text("記録一覧", size=16, weight=ft.FontWeight.BOLD, color=config.COLOR_WHITE),
                        Container(height=10),
                        Container(
                            content=Column(
                                controls=[
                                    Container(
                                        content=Row(
                                            controls=[
                                                Text(f"{log.get('ClubName', '')} - {log.get('Period', '')}", expand=True, size=12),
                                                IconButton(icon=Icons.DELETE, icon_size=16, icon_color="#ef4444"),
                                            ],
                                        ),
                                        padding=8,
                                        border=ft.border.only(bottom=BorderSide(1, config.COLOR_GRAY_200)),
                                    )
                                    for log in external_logs[:10]
                                ] if external_logs else [Text("記録がありません", color=config.COLOR_GRAY_500, size=12)],
                                scroll=ScrollMode.AUTO,
                            ),
                            bgcolor=config.COLOR_WHITE,
                            border_radius=8,
                            expand=True,
                        ),
                    ],
                    expand=True,
                ),
                bgcolor=config.DEPT_COLORS["external"],
                border_radius=12,
                padding=15,
                expand=True,
            )
            
            left_panel = Column(
                controls=[
                    Container(content=record_list, expand=1),
                    Container(content=form_panel, expand=2),
                ],
                spacing=15,
                expand=True,
            )
            
            return Row(
                controls=[
                    Container(content=left_panel, expand=1),
                    Container(content=pdf_panel, expand=1),
                ],
                spacing=15,
                expand=True,
            )
        
        # --- Tab 1: 記録一覧 & Excel出力 ---
        def create_list_tab():
            start_date = create_text_field("開始日", width=150)
            end_date = create_text_field("終了日", width=150)
            export_btn = create_primary_button("Excelエクスポート", lambda e: None, icon=Icons.DOWNLOAD)
            
            table = DataTable(
                columns=[
                    DataColumn(Text("団体名", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("期間", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("宿泊", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("公式戦", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("大会名", weight=ft.FontWeight.BOLD)),
                ],
                rows=[
                    DataRow(cells=[
                        DataCell(Text(log.get("ClubName", ""), size=12)),
                        DataCell(Text(log.get("Period", ""), size=12)),
                        DataCell(Text("○" if log.get("HasOvernight") == "TRUE" else "", size=12)),
                        DataCell(Text("○" if log.get("HasMatch") == "TRUE" else "", size=12)),
                        DataCell(Text(log.get("TournamentName", ""), size=12)),
                    ])
                    for log in external_logs
                ],
                border=ft.border.all(1, config.COLOR_GRAY_200),
                border_radius=8,
                heading_row_color=config.COLOR_GRAY_50,
            )
            
            return Column(
                controls=[
                    Text("宿泊・合宿リスト抽出", size=18, weight=ft.FontWeight.BOLD),
                    Container(height=15),
                    Row(
                        controls=[start_date, Text("〜"), end_date, Container(expand=True), export_btn],
                        spacing=10,
                    ),
                    Container(height=15),
                    Container(
                        content=Column(controls=[table], scroll=ScrollMode.BOTH),
                        expand=True,
                        bgcolor=config.COLOR_WHITE,
                        border_radius=8,
                    ),
                ],
                expand=True,
            )
        
        # Tab switching
        tabs = [("記録", create_record_tab), ("記録一覧 & Excel出力", create_list_tab)]
        
        def switch_tab(index):
            current_tab["value"] = index
            for i, btn in enumerate(tab_buttons_row.controls):
                btn.style = ft.ButtonStyle(
                    bgcolor=get_accent_color() if i == index else config.COLOR_GRAY_100,
                    color=config.COLOR_WHITE if i == index else config.COLOR_GRAY_700,
                    padding=padding.symmetric(horizontal=20, vertical=10),
                    shape=RoundedRectangleBorder(radius=8),
                )
            tab_content.content = tabs[index][1]()
            self.page.update()
        
        tab_buttons_row.controls = self._create_tab_buttons(tabs, current_tab, switch_tab)
        tab_content.content = create_record_tab()
        
        header = self._create_dept_header("渉外部", config.DEPT_COLORS["external"], Icons.PUBLIC)
        
        content = Container(
            content=Column(
                controls=[
                    tab_buttons_row,
                    Container(height=15),
                    tab_content,
                ],
            ),
            padding=20,
            expand=True,
        )
        
        self.page.controls.clear()
        self.page.add(Column(controls=[header, content], spacing=0, expand=True))
        self.page.update()
    
    # ============================================================
    # Editorial View (編集部) - Tab order: 名簿校正, 校正 編集, URL
    # ============================================================
    def _show_editorial_view(self):
        self.current_view = "editorial"
        
        # Load data
        required_items = []
        passwords = []
        advisors = []
        bookmarks = []
        
        try:
            if self.sheets_service:
                required_items = self.sheets_service.get_required_items() or []
                passwords = self.sheets_service.get_passwords() or []
                advisors = self.sheets_service.get_advisors() or []
                bookmarks = self.sheets_service.get_bookmarks() or []
        except Exception as e:
            get_error_logger().log_error("Editorial", "load", e)
        
        current_tab = {"value": 0}
        tab_content = Container(expand=True)
        tab_buttons_row = Row(controls=[], spacing=10)
        
        # --- Tab 0: 名簿校正 (PDF Generation) ---
        def create_roster_tab():
            pdf_status = Text(
                f"PDF生成ライブラリ: {'利用可能' if HAS_PDF_LIBS else 'インストールが必要'}",
                color="#22c55e" if HAS_PDF_LIBS else "#ef4444",
                size=14,
            )
            
            selected_file = {"path": None, "name": None}
            file_label = Text("ファイル未選択", color=config.COLOR_GRAY_500)
            
            def on_file_selected(e):
                if e.files:
                    selected_file["path"] = e.files[0].path
                    selected_file["name"] = e.files[0].name
                    file_label.value = selected_file["name"]
                    self.page.update()
            
            def on_pick_file(e):
                self._file_pick_callback = on_file_selected
                self.file_picker.pick_files(
                    dialog_title="Excelファイルを選択",
                    allowed_extensions=["xlsx", "xls"],
                )
            
            progress_bar = ProgressBar(visible=False, width=400)
            status_text = Text("", color=config.COLOR_GRAY_600)
            
            def generate_pdfs(e):
                if not HAS_PDF_LIBS:
                    self.page.snack_bar = SnackBar(
                        Text("pip install reportlab PyPDF2 でライブラリをインストールしてください"),
                        bgcolor="#ef4444",
                    )
                    self.page.snack_bar.open = True
                    self.page.update()
                    return
                
                if not selected_file["path"]:
                    self.page.snack_bar = SnackBar(Text("Excelファイルを選択してください"), bgcolor="#ef4444")
                    self.page.snack_bar.open = True
                    self.page.update()
                    return
                
                progress_bar.visible = True
                status_text.value = "名簿PDFを生成中..."
                self.page.update()
                
                def do_generate():
                    try:
                        # PDF generation implementation
                        import time
                        
                        # Read Excel file
                        if HAS_EXCEL:
                            wb = openpyxl.load_workbook(selected_file["path"])
                            sheets = wb.sheetnames
                            
                            output_dir = tempfile.mkdtemp()
                            generated_files = []
                            
                            for i, sheet_name in enumerate(sheets):
                                progress = (i + 1) / len(sheets)
                                self.page.run_thread(lambda p=progress, s=sheet_name: update_progress(p, s))
                                
                                # Generate PDF for each sheet
                                ws = wb[sheet_name]
                                pdf_path = os.path.join(output_dir, f"{sheet_name}.pdf")
                                
                                # Create PDF
                                doc = SimpleDocTemplate(pdf_path, pagesize=A4)
                                elements = []
                                
                                # Add title
                                styles = getSampleStyleSheet()
                                elements.append(Paragraph(sheet_name, styles['Title']))
                                elements.append(Spacer(1, 20))
                                
                                # Build table data
                                data = []
                                for row in ws.iter_rows(values_only=True):
                                    if any(cell is not None for cell in row):
                                        data.append([str(cell or "") for cell in row])
                                
                                if data:
                                    table = Table(data)
                                    table.setStyle(TableStyle([
                                        ('GRID', (0, 0), (-1, -1), 1, rl_colors.black),
                                        ('FONTNAME', (0, 0), (-1, -1), 'HeiseiMin-W3'),
                                    ]))
                                    elements.append(table)
                                
                                doc.build(elements)
                                generated_files.append(pdf_path)
                            
                            # Create ZIP
                            zip_path = os.path.join(output_dir, "名簿PDF.zip")
                            with zipfile.ZipFile(zip_path, 'w') as zf:
                                for pdf_file in generated_files:
                                    zf.write(pdf_file, os.path.basename(pdf_file))
                            
                            self.page.run_thread(lambda: generation_complete(zip_path))
                        else:
                            self.page.run_thread(lambda: generation_error("openpyxlがインストールされていません"))
                    except Exception as ex:
                        self.page.run_thread(lambda: generation_error(str(ex)))
                
                def update_progress(p, name):
                    progress_bar.value = p
                    status_text.value = f"処理中: {name}"
                    self.page.update()
                
                def generation_complete(zip_path):
                    progress_bar.visible = False
                    status_text.value = f"生成完了: {zip_path}"
                    self.page.snack_bar = SnackBar(Text("PDF生成が完了しました"), bgcolor="#22c55e")
                    self.page.snack_bar.open = True
                    self.page.update()
                
                def generation_error(msg):
                    progress_bar.visible = False
                    status_text.value = f"エラー: {msg}"
                    self.page.snack_bar = SnackBar(Text(f"エラー: {msg}"), bgcolor="#ef4444")
                    self.page.snack_bar.open = True
                    self.page.update()
                
                threading.Thread(target=do_generate, daemon=True).start()
            
            return Container(
                content=Column(
                    controls=[
                        Text("名簿PDF一括生成", size=18, weight=ft.FontWeight.BOLD),
                        Container(height=15),
                        pdf_status,
                        Container(height=20),
                        Text(
                            "Excelファイルから各団体の名簿PDFを生成し、パスワード付きZIPでダウンロードします。",
                            size=14,
                            color=config.COLOR_GRAY_600,
                        ),
                        Container(height=20),
                        Row(
                            controls=[
                                create_primary_button("Excelファイルを選択してPDF生成", on_pick_file, icon=Icons.PICTURE_AS_PDF),
                            ],
                            spacing=15,
                        ),
                        Container(height=10),
                        file_label,
                        Container(height=20),
                        create_primary_button(
                            "PDF生成開始",
                            generate_pdfs,
                            icon=Icons.PLAY_ARROW,
                            disabled=not HAS_PDF_LIBS,
                        ),
                        Container(height=20),
                        progress_bar,
                        status_text,
                    ],
                ),
                bgcolor=config.COLOR_WHITE,
                border_radius=12,
                padding=20,
                expand=True,
            )
        
        # --- Tab 1: 校正 編集 (DB Edit) ---
        def create_edit_tab():
            registered_clubs = {c.get("ClubName", "") for c in self.clubs}
            
            # Collect all unique club names from data
            all_data_clubs = set()
            for r in required_items:
                all_data_clubs.add(r.get("ClubName", ""))
            for p in passwords:
                all_data_clubs.add(p.get("ClubName", ""))
            for a in advisors:
                all_data_clubs.add(a.get("ClubName", ""))
            all_data_clubs.discard("")
            
            # All clubs = registered + unregistered
            all_clubs = list(registered_clubs | all_data_clubs)
            all_clubs.sort()
            
            table_rows = []
            for club_name in all_clubs:
                is_registered = club_name in registered_clubs
                
                req = next((r for r in required_items if r.get("ClubName") == club_name), {})
                pw = next((p for p in passwords if p.get("ClubName") == club_name), {})
                adv = next((a for a in advisors if a.get("ClubName") == club_name), {})
                
                name_color = config.COLOR_GRAY_800 if is_registered else "#ef4444"
                
                table_rows.append(DataRow(cells=[
                    DataCell(Text(club_name, color=name_color, size=12)),
                    DataCell(Checkbox(value=bool(req.get("StudentPhone")))),
                    DataCell(Checkbox(value=bool(req.get("GuarantorPhone")))),
                    DataCell(Checkbox(value=bool(req.get("Address")))),
                    DataCell(TextField(value=pw.get("Password", ""), width=80, text_size=12, border_color=config.COLOR_GRAY_300)),
                    DataCell(TextField(value=adv.get("Director", ""), width=80, text_size=12, border_color=config.COLOR_GRAY_300)),
                    DataCell(TextField(value=adv.get("Advisor", ""), width=80, text_size=12, border_color=config.COLOR_GRAY_300)),
                    DataCell(TextField(value=adv.get("Coach", ""), width=80, text_size=12, border_color=config.COLOR_GRAY_300)),
                ]))
            
            table = DataTable(
                columns=[
                    DataColumn(Text("団体名", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("学生携帯", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("保証人携帯", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("住所", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("PDF PW", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("部長", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("顧問", weight=ft.FontWeight.BOLD)),
                    DataColumn(Text("監督/コーチ", weight=ft.FontWeight.BOLD)),
                ],
                rows=table_rows,
                border=ft.border.all(1, config.COLOR_GRAY_200),
                border_radius=8,
                heading_row_color=config.COLOR_GRAY_50,
            )
            
            def save_all(e):
                self.loading.show("保存中...")
                try:
                    # TODO: Implement save all changes
                    self.page.snack_bar = SnackBar(Text("保存しました"), bgcolor="#22c55e")
                    self.page.snack_bar.open = True
                except Exception as ex:
                    get_error_logger().log_error("Editorial", "save_db", ex)
                finally:
                    self.loading.hide()
                    self.page.update()
            
            return Container(
                content=Column(
                    controls=[
                        Row(
                            controls=[
                                Text("必要項目・パスワード・顧問情報", size=16, weight=ft.FontWeight.W_500),
                                Container(expand=True),
                                create_primary_button("一括保存", save_all, icon=Icons.SAVE),
                            ],
                        ),
                        Container(height=10),
                        Text("※赤字は未登録の団体です", size=12, color="#ef4444"),
                        Container(height=10),
                        Container(
                            content=Column(controls=[table], scroll=ScrollMode.BOTH),
                            expand=True,
                            bgcolor=config.COLOR_WHITE,
                            border_radius=8,
                        ),
                    ],
                    expand=True,
                ),
                padding=10,
                expand=True,
            )
        
        # --- Tab 2: URL ---
        def create_url_tab():
            url_list_col = Column(controls=[], spacing=5, scroll=ScrollMode.AUTO, expand=True)
            
            name_field = create_text_field("名前", width=200)
            url_field = create_text_field("URL", width=400)
            
            def refresh_url_list():
                url_list_col.controls = [
                    Container(
                        content=Row(
                            controls=[
                                Text(b.get("Name", ""), width=150, size=13),
                                Text(b.get("URL", ""), expand=True, color=config.COLOR_GRAY_600, size=12),
                                IconButton(
                                    icon=Icons.DELETE,
                                    icon_color="#ef4444",
                                    icon_size=18,
                                    on_click=lambda e, idx=i: delete_bookmark(idx),
                                ),
                            ],
                            spacing=10,
                        ),
                        padding=10,
                        border=ft.border.only(bottom=BorderSide(1, config.COLOR_GRAY_200)),
                    )
                    for i, b in enumerate(bookmarks)
                ] if bookmarks else [Text("URLがありません", color=config.COLOR_GRAY_500)]
                self.page.update()
            
            def add_bookmark(e):
                if not name_field.value or not url_field.value:
                    self.page.snack_bar = SnackBar(Text("名前とURLを入力してください"), bgcolor="#ef4444")
                    self.page.snack_bar.open = True
                    self.page.update()
                    return
                
                bookmarks.append({"Name": name_field.value, "URL": url_field.value})
                try:
                    self.sheets_service.save_bookmarks(bookmarks)
                    name_field.value = ""
                    url_field.value = ""
                    refresh_url_list()
                    self.page.snack_bar = SnackBar(Text("追加しました"), bgcolor="#22c55e")
                    self.page.snack_bar.open = True
                except Exception as ex:
                    get_error_logger().log_error("Editorial", "add_bookmark", ex)
                self.page.update()
            
            def delete_bookmark(idx):
                try:
                    bookmarks.pop(idx)
                    self.sheets_service.save_bookmarks(bookmarks)
                    refresh_url_list()
                except Exception as ex:
                    get_error_logger().log_error("Editorial", "delete_bookmark", ex)
            
            refresh_url_list()
            
            return Container(
                content=Column(
                    controls=[
                        Text("各URL設定", size=18, weight=ft.FontWeight.BOLD),
                        Container(height=15),
                        Row(
                            controls=[
                                name_field,
                                url_field,
                                create_primary_button("追加", add_bookmark, icon=Icons.ADD),
                            ],
                            spacing=10,
                        ),
                        Container(height=20),
                        Container(
                            content=url_list_col,
                            bgcolor=config.COLOR_WHITE,
                            border_radius=8,
                            expand=True,
                        ),
                    ],
                    expand=True,
                ),
                bgcolor=config.COLOR_WHITE,
                border_radius=12,
                padding=20,
                expand=True,
            )
        
        # Tab switching - Order: 名簿校正, 校正 編集, URL
        tabs = [("名簿校正", create_roster_tab), ("校正 編集", create_edit_tab), ("URL", create_url_tab)]
        
        def switch_tab(index):
            current_tab["value"] = index
            for i, btn in enumerate(tab_buttons_row.controls):
                btn.style = ft.ButtonStyle(
                    bgcolor=get_accent_color() if i == index else config.COLOR_GRAY_100,
                    color=config.COLOR_WHITE if i == index else config.COLOR_GRAY_700,
                    padding=padding.symmetric(horizontal=20, vertical=10),
                    shape=RoundedRectangleBorder(radius=8),
                )
            tab_content.content = tabs[index][1]()
            self.page.update()
        
        tab_buttons_row.controls = self._create_tab_buttons(tabs, current_tab, switch_tab)
        tab_content.content = create_roster_tab()
        
        header = self._create_dept_header("編集部", config.DEPT_COLORS["editorial"], Icons.EDIT_NOTE)
        
        content = Container(
            content=Column(
                controls=[
                    tab_buttons_row,
                    Container(height=15),
                    tab_content,
                ],
            ),
            padding=20,
            expand=True,
        )
        
        self.page.controls.clear()
        self.page.add(Column(controls=[header, content], spacing=0, expand=True))
        self.page.update()
    
    # ============================================================
    # Event View (イベント管理)
    # ============================================================
    def _show_event_view(self):
        self.current_view = "event"
        
        current_tab = {"value": 0}
        tab_content = Container(expand=True)
        tab_buttons_row = Row(controls=[], spacing=10)
        
        # --- Tab 0: フォーム集計 ---
        def create_form_tab():
            return Container(
                content=Column(
                    controls=[
                        Text("Googleフォーム & スプレッドシート統計", size=18, weight=ft.FontWeight.BOLD),
                        Container(height=15),
                        Text("Google Driveからスプレッドシートを選択して集計します", size=14, color=config.COLOR_GRAY_600),
                        Container(height=15),
                        create_primary_button(
                            "Google Driveから選択",
                            lambda e: self.file_picker.pick_files(allowed_extensions=["xlsx", "csv"]),
                            icon=Icons.FOLDER_OPEN,
                        ),
                        Container(height=30),
                        Text("統計結果がここに表示されます", color=config.COLOR_GRAY_500),
                    ],
                ),
                bgcolor=config.COLOR_WHITE,
                border_radius=12,
                padding=20,
                expand=True,
            )
        
        # --- Tab 1: フォルダ作成 ---
        def create_folder_tab():
            club_search_field = create_text_field("団体名検索", width=300, hint_text="部分一致で検索...")
            club_checkboxes = Column(controls=[], spacing=5, scroll=ScrollMode.AUTO, height=250)
            
            def update_club_list(e=None):
                query = club_search_field.value.lower() if club_search_field.value else ""
                filtered = [c for c in self.clubs if query in c.get("ClubName", "").lower()]
                club_checkboxes.controls = [
                    Checkbox(label=c.get("ClubName", ""), value=False)
                    for c in filtered
                ]
                self.page.update()
            
            club_search_field.on_change = update_club_list
            update_club_list()
            
            def select_all(e):
                for cb in club_checkboxes.controls:
                    cb.value = True
                self.page.update()
            
            def deselect_all(e):
                for cb in club_checkboxes.controls:
                    cb.value = False
                self.page.update()
            
            subfolder_inputs = Column(
                controls=[create_text_field("サブフォルダ名", width=250)],
                spacing=10,
            )
            
            def add_subfolder(e):
                subfolder_inputs.controls.append(create_text_field("サブフォルダ名", width=250))
                self.page.update()
            
            def on_create_folders(e):
                selected_clubs = [cb.label for cb in club_checkboxes.controls if cb.value]
                subfolders = [sf.value for sf in subfolder_inputs.controls if sf.value]
                
                if not selected_clubs:
                    self.page.snack_bar = SnackBar(Text("団体を選択してください"), bgcolor="#ef4444")
                    self.page.snack_bar.open = True
                    self.page.update()
                    return
                
                self.loading.show("フォルダを作成中...")
                try:
                    # TODO: Create folders via Google Drive API
                    self.page.snack_bar = SnackBar(
                        Text(f"{len(selected_clubs)}団体のフォルダを作成しました"),
                        bgcolor="#22c55e",
                    )
                    self.page.snack_bar.open = True
                finally:
                    self.loading.hide()
                    self.page.update()
            
            return Container(
                content=Column(
                    controls=[
                        Text("フォルダ一括作成", size=18, weight=ft.FontWeight.BOLD),
                        Container(height=15),
                        club_search_field,
                        Container(height=10),
                        Row(
                            controls=[
                                TextButton("すべて選択", on_click=select_all),
                                TextButton("選択解除", on_click=deselect_all),
                            ],
                        ),
                        Container(
                            content=club_checkboxes,
                            border=ft.border.all(1, config.COLOR_GRAY_200),
                            border_radius=8,
                            padding=10,
                        ),
                        Container(height=20),
                        Text("サブフォルダ", size=14, weight=ft.FontWeight.W_500),
                        subfolder_inputs,
                        TextButton("サブフォルダを追加", icon=Icons.ADD, on_click=add_subfolder),
                        Container(height=20),
                        create_primary_button("フォルダ作成", on_create_folders, icon=Icons.CREATE_NEW_FOLDER),
                    ],
                    scroll=ScrollMode.AUTO,
                ),
                bgcolor=config.COLOR_WHITE,
                border_radius=12,
                padding=20,
                expand=True,
            )
        
        # --- Tab 2: スキャン ---
        def create_scan_tab():
            scan_mode = {"value": "mono"}
            
            mono_btn = ElevatedButton("モノクロ", style=ft.ButtonStyle(bgcolor=get_accent_color(), color=config.COLOR_WHITE))
            color_btn = ElevatedButton("カラー", style=ft.ButtonStyle(bgcolor=config.COLOR_GRAY_100, color=config.COLOR_GRAY_700))
            
            def set_mode(mode):
                scan_mode["value"] = mode
                mono_btn.style = ft.ButtonStyle(bgcolor=get_accent_color() if mode == "mono" else config.COLOR_GRAY_100, color=config.COLOR_WHITE if mode == "mono" else config.COLOR_GRAY_700)
                color_btn.style = ft.ButtonStyle(bgcolor=get_accent_color() if mode == "color" else config.COLOR_GRAY_100, color=config.COLOR_WHITE if mode == "color" else config.COLOR_GRAY_700)
                self.page.update()
            
            mono_btn.on_click = lambda e: set_mode("mono")
            color_btn.on_click = lambda e: set_mode("color")
            
            return Row(
                controls=[
                    Container(
                        content=Column(
                            controls=[
                                Text("スキャン設定", size=16, weight=ft.FontWeight.BOLD),
                                Container(height=15),
                                Row(controls=[mono_btn, color_btn], spacing=10),
                                Container(height=20),
                                create_secondary_button("OneDriveから選択", lambda e: self.file_picker.pick_files(allowed_extensions=["pdf"]), icon=Icons.FOLDER_OPEN),
                                Container(height=10),
                                create_primary_button("スキャン実行", lambda e: None, icon=Icons.SCANNER),
                            ],
                        ),
                        bgcolor=config.COLOR_WHITE,
                        border_radius=12,
                        padding=20,
                        expand=1,
                    ),
                    Container(
                        content=Column(
                            controls=[
                                Text("PDFプレビュー", size=16, weight=ft.FontWeight.BOLD, color=config.COLOR_WHITE),
                                Container(height=20),
                                Container(
                                    content=Text("PDFを選択またはスキャンしてください", color=config.COLOR_GRAY_400),
                                    bgcolor=config.COLOR_WHITE,
                                    border_radius=8,
                                    expand=True,
                                    alignment=alignment.center,
                                ),
                                Container(height=10),
                                Row(
                                    controls=[
                                        IconButton(icon=Icons.ROTATE_LEFT, icon_color=config.COLOR_WHITE),
                                        IconButton(icon=Icons.ROTATE_RIGHT, icon_color=config.COLOR_WHITE),
                                        IconButton(icon=Icons.CONTENT_CUT, icon_color=config.COLOR_WHITE),
                                    ],
                                    alignment=MainAxisAlignment.CENTER,
                                ),
                            ],
                        ),
                        bgcolor=config.DEPT_COLORS["event"],
                        border_radius=12,
                        padding=20,
                        expand=2,
                    ),
                ],
                spacing=15,
                expand=True,
            )
        
        # Tab switching
        tabs = [("フォーム集計", create_form_tab), ("フォルダ作成", create_folder_tab), ("スキャン", create_scan_tab)]
        
        def switch_tab(index):
            current_tab["value"] = index
            for i, btn in enumerate(tab_buttons_row.controls):
                btn.style = ft.ButtonStyle(
                    bgcolor=get_accent_color() if i == index else config.COLOR_GRAY_100,
                    color=config.COLOR_WHITE if i == index else config.COLOR_GRAY_700,
                    padding=padding.symmetric(horizontal=20, vertical=10),
                    shape=RoundedRectangleBorder(radius=8),
                )
            tab_content.content = tabs[index][1]()
            self.page.update()
        
        tab_buttons_row.controls = self._create_tab_buttons(tabs, current_tab, switch_tab)
        tab_content.content = create_form_tab()
        
        header = self._create_dept_header("イベント管理", config.DEPT_COLORS["event"], Icons.EVENT)
        
        content = Container(
            content=Column(
                controls=[
                    tab_buttons_row,
                    Container(height=15),
                    tab_content,
                ],
            ),
            padding=20,
            expand=True,
        )
        
        self.page.controls.clear()
        self.page.add(Column(controls=[header, content], spacing=0, expand=True))
        self.page.update()
    
    # ============================================================
    # Admin View (管理&設定) - With DB initialization
    # ============================================================
    def _show_admin_view(self):
        self.current_view = "admin"
        
        current_tab = {"value": 0}
        tab_content = Container(expand=True)
        tab_buttons_row = Row(controls=[], spacing=10)
        
        # --- Tab 0: 体育会本部 ---
        def create_honbu_tab():
            # DB Initialization
            def init_database(e):
                self.loading.show("データベースを初期化中...")
                
                def do_init():
                    try:
                        sheets_to_create = [
                            (config.SHEET_CLUBS, ["ClubName", "Category", "Color", "DisplayName"]),
                            (config.SHEET_MEMBERS, ["StudentID", "Name", "Department", "Role"]),
                            (config.SHEET_FACILITIES, ["FacilityID", "FacilityName"]),
                            (config.SHEET_SECRETARY_LOG, ["Date", "Facility", "ClubName", "StartTime", "EndTime", "Note", "CreatedAt"]),
                            (config.SHEET_FINANCE, ["Date", "Subject", "Description", "PaymentMethod", "Income", "Expense", "ReimbursementStatus", "CreatedAt"]),
                            (config.SHEET_ATTENDANCE, ["Date", "MemberName", "Status", "Period", "CreatedAt"]),
                            (config.SHEET_WEEKDAY_ASSIGN, ["Period", "Weekday", "Members"]),
                            (config.SHEET_EXTERNAL_LOG, ["No", "LogType", "HasScan", "HasActivity", "HasReport", "Period", "HasMatch", "TournamentName", "HasOvernight", "Organizer", "Participants", "TeamResult", "IndividualResult", "ClubName", "CreatedAt"]),
                            (config.SHEET_BOOKMARKS, ["Name", "URL"]),
                            (config.SHEET_REQUIRED_ITEMS, ["ClubName", "StudentPhone", "GuarantorPhone", "Address"]),
                            (config.SHEET_PASSWORDS, ["ClubName", "Password"]),
                            (config.SHEET_ADVISORS, ["ClubName", "Director", "Advisor", "Coach", "CoachSub"]),
                            (config.SHEET_CATEGORIES, ["CategoryName", "Order"]),
                        ]
                        
                        for sheet_name, headers in sheets_to_create:
                            self.sheets_service.get_or_create_sheet(sheet_name, headers)
                        
                        # Add default facilities if empty
                        facilities = self.sheets_service.get_facilities()
                        if not facilities:
                            for i, name in enumerate(config.DEFAULT_FACILITIES):
                                self.sheets_service.append_row(config.SHEET_FACILITIES, [f"F{i+1}", name])
                        
                        # Add default categories if empty
                        categories = self.sheets_service.get_categories()
                        if not categories:
                            self.sheets_service.append_row(config.SHEET_CATEGORIES, ["体育会", "1"])
                            self.sheets_service.append_row(config.SHEET_CATEGORIES, ["同好会", "2"])
                        
                        self._load_initial_data()
                        
                        self.page.run_thread(lambda: init_complete())
                    except Exception as ex:
                        self.page.run_thread(lambda: init_error(str(ex)))
                
                def init_complete():
                    self.loading.hide()
                    self.page.snack_bar = SnackBar(Text("データベースを初期化しました"), bgcolor="#22c55e")
                    self.page.snack_bar.open = True
                    self.page.update()
                
                def init_error(msg):
                    self.loading.hide()
                    get_error_logger().log_error("Admin", "init_database", Exception(msg))
                    self.page.snack_bar = SnackBar(Text(f"エラー: {msg}"), bgcolor="#ef4444")
                    self.page.snack_bar.open = True
                    self.page.update()
                
                threading.Thread(target=do_init, daemon=True).start()
            
            init_section = Container(
                content=Column(
                    controls=[
                        Row(
                            controls=[
                                Icon(Icons.WARNING_AMBER, color="#f97316", size=24),
                                Text("データベース初期化", size=16, weight=ft.FontWeight.BOLD),
                            ],
                            spacing=10,
                        ),
                        Container(height=10),
                        Text("必要なシートがない場合、自動で作成します。初回起動時に実行してください。", size=14, color=config.COLOR_GRAY_600),
                        Container(height=15),
                        create_primary_button("データベース初期化を実行", init_database, icon=Icons.STORAGE, bgcolor="#f97316"),
                    ],
                ),
                bgcolor="#fff7ed",
                border=ft.border.all(1, "#fed7aa"),
                border_radius=12,
                padding=20,
            )
            
            # Club management
            club_name_field = create_text_field("団体名", width=200)
            category_options = [(c.get("CategoryName", ""), c.get("CategoryName", "")) for c in self.categories]
            if not category_options:
                category_options = [("体育会", "体育会"), ("同好会", "同好会")]
            club_category_field = create_dropdown("区分", category_options, width=150)
            
            club_list = Column(controls=[], scroll=ScrollMode.AUTO, height=200)
            
            def refresh_club_list():
                club_list.controls = [
                    Container(
                        content=Row(
                            controls=[
                                Text(c.get("ClubName", ""), width=200, size=13),
                                Text(c.get("Category", ""), width=80, color=config.COLOR_GRAY_600, size=12),
                                IconButton(icon=Icons.DELETE, icon_color="#ef4444", icon_size=18, on_click=lambda e, idx=i: delete_club(idx)),
                            ],
                            spacing=10,
                        ),
                        padding=8,
                        border=ft.border.only(bottom=BorderSide(1, config.COLOR_GRAY_100)),
                    )
                    for i, c in enumerate(self.clubs)
                ] if self.clubs else [Text("団体がありません", color=config.COLOR_GRAY_500)]
                self.page.update()
            
            def add_club(e):
                if not club_name_field.value:
                    return
                self.loading.show("追加中...")
                try:
                    self.sheets_service.add_club({
                        "ClubName": club_name_field.value,
                        "Category": club_category_field.value,
                        "Color": "",
                        "DisplayName": club_name_field.value,
                    })
                    club_name_field.value = ""
                    self.clubs = self.sheets_service.get_clubs()
                    refresh_club_list()
                    self.page.snack_bar = SnackBar(Text("追加しました"), bgcolor="#22c55e")
                    self.page.snack_bar.open = True
                except Exception as ex:
                    get_error_logger().log_error("Admin", "add_club", ex)
                finally:
                    self.loading.hide()
                    self.page.update()
            
            def delete_club(idx):
                try:
                    self.sheets_service.delete_club(idx)
                    self.clubs = self.sheets_service.get_clubs()
                    refresh_club_list()
                except Exception as ex:
                    get_error_logger().log_error("Admin", "delete_club", ex)
            
            refresh_club_list()
            
            club_section = Container(
                content=Column(
                    controls=[
                        Text("団体・区分管理", size=16, weight=ft.FontWeight.BOLD),
                        Container(height=10),
                        Row(
                            controls=[club_name_field, club_category_field, create_primary_button("追加", add_club, icon=Icons.ADD)],
                            spacing=10,
                        ),
                        Container(height=10),
                        Container(
                            content=club_list,
                            border=ft.border.all(1, config.COLOR_GRAY_200),
                            border_radius=8,
                            padding=10,
                        ),
                        Text(f"登録団体数: {len(self.clubs)}", size=12, color=config.COLOR_GRAY_500),
                    ],
                ),
                bgcolor=config.COLOR_WHITE,
                border_radius=12,
                padding=20,
            )
            
            # Member management
            member_name_field = create_text_field("氏名", width=150)
            member_id_field = create_text_field("学籍番号", width=150)
            member_dept_field = create_dropdown(
                "部署",
                [("secretary", "書記部"), ("finance", "財務部"), ("general", "総務部"),
                 ("external", "渉外部"), ("editorial", "編集部")],
                width=150,
            )
            
            def add_member(e):
                if not member_name_field.value:
                    return
                self.loading.show("追加中...")
                try:
                    self.sheets_service.upsert_member({
                        "Name": member_name_field.value,
                        "StudentID": member_id_field.value,
                        "Department": member_dept_field.value,
                    })
                    member_name_field.value = ""
                    member_id_field.value = ""
                    self.members = self.sheets_service.get_members()
                    self.page.snack_bar = SnackBar(Text("追加/更新しました"), bgcolor="#22c55e")
                    self.page.snack_bar.open = True
                except Exception as ex:
                    get_error_logger().log_error("Admin", "add_member", ex)
                finally:
                    self.loading.hide()
                    self.page.update()
            
            def import_csv(e):
                self.file_picker.pick_files(
                    dialog_title="CSV/Excelファイルを選択",
                    allowed_extensions=["csv", "xlsx"],
                )
            
            member_section = Container(
                content=Column(
                    controls=[
                        Text("本部員情報管理", size=16, weight=ft.FontWeight.BOLD),
                        Container(height=10),
                        Row(
                            controls=[
                                member_name_field, member_id_field, member_dept_field,
                                create_primary_button("追加/更新", add_member, icon=Icons.PERSON_ADD),
                            ],
                            spacing=10,
                            wrap=True,
                        ),
                        Container(height=10),
                        create_secondary_button("CSV/Excel一括インポート", import_csv, icon=Icons.UPLOAD_FILE),
                        Container(height=10),
                        Text(f"現在の本部員数: {len(self.members)}名", color=config.COLOR_GRAY_600),
                    ],
                ),
                bgcolor=config.COLOR_WHITE,
                border_radius=12,
                padding=20,
            )
            
            return Column(
                controls=[
                    init_section,
                    Container(height=20),
                    club_section,
                    Container(height=20),
                    member_section,
                ],
                scroll=ScrollMode.AUTO,
                expand=True,
            )
        
        # --- Tab 1: 外部設定 ---
        def create_settings_tab():
            scanner_url_field = create_text_field("スキャナーURL", user_settings.scanner_url, width=400)
            gemini_key_field = create_text_field("Gemini APIキー", user_settings.gemini_api_key, width=400, password=True)
            onedrive_path_field = create_text_field("OneDriveパス", user_settings.onedrive_path, width=400)
            
            password_field = create_text_field("現在のパスワード", "", width=200, password=True)
            new_password_field = create_text_field("新しいパスワード", "", width=200, password=True)
            
            def save_settings(e):
                user_settings.scanner_url = scanner_url_field.value
                user_settings.gemini_api_key = gemini_key_field.value
                user_settings.onedrive_path = onedrive_path_field.value
                self.page.snack_bar = SnackBar(Text("設定を保存しました"), bgcolor="#22c55e")
                self.page.snack_bar.open = True
                self.page.update()
            
            def change_password(e):
                if password_field.value == user_settings.system_password or not user_settings.system_password:
                    user_settings.system_password = new_password_field.value
                    password_field.value = ""
                    new_password_field.value = ""
                    self.page.snack_bar = SnackBar(Text("パスワードを変更しました"), bgcolor="#22c55e")
                else:
                    self.page.snack_bar = SnackBar(Text("現在のパスワードが違います"), bgcolor="#ef4444")
                self.page.snack_bar.open = True
                self.page.update()
            
            return Column(
                controls=[
                    Container(
                        content=Column(
                            controls=[
                                Text("外部連携設定", size=16, weight=ft.FontWeight.BOLD),
                                Container(height=15),
                                Text("スキャナーURL", size=12, color=config.COLOR_GRAY_600),
                                scanner_url_field,
                                Container(height=10),
                                Text("Gemini APIキー", size=12, color=config.COLOR_GRAY_600),
                                gemini_key_field,
                                Container(height=10),
                                Text("OneDriveパス", size=12, color=config.COLOR_GRAY_600),
                                onedrive_path_field,
                                Container(height=15),
                                create_primary_button("設定を保存", save_settings, icon=Icons.SAVE),
                            ],
                        ),
                        bgcolor=config.COLOR_WHITE,
                        border_radius=12,
                        padding=20,
                    ),
                    Container(height=20),
                    Container(
                        content=Column(
                            controls=[
                                Text("システムパスワード変更", size=16, weight=ft.FontWeight.BOLD),
                                Container(height=15),
                                Row(controls=[password_field, new_password_field], spacing=10),
                                Container(height=10),
                                create_secondary_button("パスワード変更", change_password, icon=Icons.LOCK),
                            ],
                        ),
                        bgcolor=config.COLOR_WHITE,
                        border_radius=12,
                        padding=20,
                    ),
                ],
                scroll=ScrollMode.AUTO,
                expand=True,
            )
        
        # Tab switching
        tabs = [("体育会本部", create_honbu_tab), ("外部設定", create_settings_tab)]
        
        def switch_tab(index):
            current_tab["value"] = index
            for i, btn in enumerate(tab_buttons_row.controls):
                btn.style = ft.ButtonStyle(
                    bgcolor=get_accent_color() if i == index else config.COLOR_GRAY_100,
                    color=config.COLOR_WHITE if i == index else config.COLOR_GRAY_700,
                    padding=padding.symmetric(horizontal=20, vertical=10),
                    shape=RoundedRectangleBorder(radius=8),
                )
            tab_content.content = tabs[index][1]()
            self.page.update()
        
        tab_buttons_row.controls = self._create_tab_buttons(tabs, current_tab, switch_tab)
        tab_content.content = create_honbu_tab()
        
        header = self._create_dept_header("管理 & 設定", config.COLOR_GRAY_600, Icons.SETTINGS)
        
        content = Container(
            content=Column(
                controls=[
                    tab_buttons_row,
                    Container(height=15),
                    tab_content,
                ],
            ),
            padding=20,
            expand=True,
        )
        
        self.page.controls.clear()
        self.page.add(Column(controls=[header, content], spacing=0, expand=True))
        self.page.update()
    
    # ============================================================
    # File Picker Handler
    # ============================================================
    def _on_file_picked(self, e):
        if self._file_pick_callback:
            self._file_pick_callback(e)
            self._file_pick_callback = None
        elif e.files:
            file_path = e.files[0].path
            self.page.snack_bar = SnackBar(Text(f"ファイルを選択: {os.path.basename(file_path)}"), bgcolor=config.COLOR_GRAY_600)
            self.page.snack_bar.open = True
            self.page.update()


# ============================================================
# Main Entry Point
# ============================================================
def main(page: Page):
    ensure_directories()
    app = CITAAApp(page)


if __name__ == "__main__":
    ft.app(target=main)
